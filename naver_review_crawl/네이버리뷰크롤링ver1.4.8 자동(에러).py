import sys
import tkinter as tk
from tkinter import ttk, filedialog, scrolledtext, messagebox
import threading
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import random
import os
import requests
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment
import subprocess
from PIL import Image as PILImage
import io

class CrawlerThread(threading.Thread):
    def __init__(self, url, pages, folder_path, sort_option, chunk_size=50, callback=None):
        threading.Thread.__init__(self)
        self.url = url
        self.pages = pages
        self.folder_path = folder_path
        self.sort_option = sort_option
        self.chunk_size = chunk_size
        self.callback = callback

    def run(self):
        self.crawl_reviews()

    def update_log(self, message):
        if self.callback:
            self.callback(message)

    def crawl_reviews(self):
        self.update_log("=" * 50)
        self.update_log("🚀 크롤링을 시작합니다...")
        self.update_log("=" * 50)

        # Selenium 설정
        self.update_log("📋 Selenium 옵션을 설정합니다...")
        options = Options()
        options.add_experimental_option("detach", True)
        options.add_argument("disable-blink-features=AutomationControlled")
        options.add_experimental_option("excludeSwitches", ["enable-logging"])
        self.update_log("✅ Selenium 옵션 설정 완료")

        self.update_log("🌐 Chrome 드라이버를 시작합니다...")
        driver = webdriver.Chrome(options=options)
        self.update_log("✅ Chrome 드라이버 시작 완료")

        # 먼저 네이버로 이동
        self.update_log("🏠 네이버 메인 페이지로 이동합니다...")
        driver.get("https://www.naver.com")
        self.update_log("✅ 네이버에 접속했습니다.")
        time.sleep(random.uniform(2, 5))

        # 변수 정의
        self.update_log("⚙️ 크롤링 설정을 초기화합니다...")
        url = self.url
        last_page_to_crawl = int(self.pages) if self.pages.isdigit() else self.pages
        sleep_time = 2
        self.update_log(f"📊 설정된 페이지 수: {last_page_to_crawl}")
        self.update_log(f"🔗 대상 URL: {url}")
        self.update_log(f"⏱️ 대기 시간: {sleep_time}초")

        # 네이버 (새로운 구조)
        product_review_button_selector = 'a[data-name="REVIEW"]'
        new_review_button_selector = 'a[data-name="REVIEW"]'  # 리뷰 탭 클릭 후 정렬 옵션 확인 필요
        bad_review_button_selector = 'a[data-name="REVIEW"]'  # 리뷰 탭 클릭 후 정렬 옵션 확인 필요
        review_selector = "#REVIEW > div > div._2LvIMaBiIO > div._2g7PKvqCKe > ul > li"  # 리뷰 탭 클릭 후 확인 필요

        # 네이버 접속 후 5초 대기
        self.update_log("⏳ 네이버 접속 후 5초 대기합니다...")
        for i in range(5, 0, -1):
            self.update_log(f"⏰ 대기 중: {i}초")
            time.sleep(1)
        self.update_log("✅ 네이버 대기 완료!")

        # 검색 페이지로 이동
        search_url = "https://search.naver.com/search.naver?where=nexearch&sm=top_hty&fbm=0&ie=utf8&query=%EC%98%A4%ED%8E%98%EB%9D%BC%EA%B8%80%EB%9D%BC%EC%8A%A4&ackey=ayy89dsf"
        self.update_log("🔍 네이버 검색 페이지로 이동합니다...")
        driver.get(search_url)
        self.update_log("✅ 검색 페이지에 접속했습니다.")
        time.sleep(random.uniform(2, 5))

        # 웹 페이지 열기
        self.update_log("🛍️ 상품 페이지로 이동합니다...")
        driver.get(url)
        self.update_log("✅ 상품 페이지에 접속했습니다.")
        
        # 15초 카운트다운
        self.update_log("⏳ 페이지 로딩을 위해 15초 대기합니다...")
        for i in range(15, 0, -1):
            self.update_log(f"⏰ 카운트다운: {i}초")
            time.sleep(1)
        self.update_log("✅ 카운트다운 완료! 다음 액션을 진행합니다.")
        
        time.sleep(sleep_time)

        # 페이지 끝까지 스크롤
        self.update_log("📜 페이지 하단으로 스크롤합니다...")
        try:
            driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            self.update_log("✅ 페이지 하단으로 스크롤했습니다.")
        except Exception as e:
            self.update_log(f"❌ 스크롤 실패: {e}")
            self.close_alert_if_present(driver)

        time.sleep(sleep_time)

        # 리뷰 탭 클릭 (JavaScript 사용)
        self.update_log("🔍 리뷰 탭을 찾는 중...")
        try:
            product_review_button = driver.find_element(By.CSS_SELECTOR, product_review_button_selector)
            self.update_log("✅ 리뷰 탭을 찾았습니다.")
            # JavaScript를 사용해서 클릭 (다른 요소에 가려진 경우 해결)
            driver.execute_script("arguments[0].click();", product_review_button)
            self.update_log("✅ 리뷰 탭을 클릭했습니다.")
        except Exception as e:
            self.update_log(f"❌ 리뷰 탭 클릭 실패: {e}")
            # 대안: JavaScript로 직접 클릭 시도
            self.update_log("🔄 JavaScript로 직접 클릭을 시도합니다...")
            try:
                driver.execute_script("document.querySelector('a[data-name=\"REVIEW\"]').click();")
                self.update_log("✅ JavaScript로 리뷰 탭을 클릭했습니다.")
            except Exception as e2:
                self.update_log(f"❌ JavaScript 클릭도 실패: {e2}")
            self.close_alert_if_present(driver)

        time.sleep(3)

        # 리뷰 탭이 제대로 클릭되었는지 확인
        self.update_log("🔍 리뷰 탭 활성화 상태를 확인합니다...")
        try:
            # 리뷰 탭이 활성화되었는지 확인 (aria-current="true" 또는 클래스 변경 확인)
            review_tab = driver.find_element(By.CSS_SELECTOR, 'a[data-name="REVIEW"]')
            if review_tab.get_attribute('aria-current') == 'true':
                self.update_log("✅ 리뷰 탭이 성공적으로 활성화되었습니다.")
            else:
                self.update_log("⚠️ 리뷰 탭 활성화 상태를 확인할 수 없습니다.")
        except Exception as e:
            self.update_log(f"❌ 리뷰 탭 상태 확인 실패: {e}")

        # 리뷰 탭 클릭 후 정렬 옵션 찾기 및 클릭
        self.update_log("🔧 정렬 옵션을 설정합니다...")
        try:
            # 정렬 옵션들을 찾기 위해 잠시 대기
            self.update_log("⏳ 정렬 옵션 로딩을 위해 3초 대기...")
            time.sleep(3)
            
            # 정렬 옵션에 따른 버튼 클릭 (새로운 구조)
            if self.sort_option == "latest":
                self.update_log("📅 최신순 정렬을 시도합니다...")
                # 최신순 정렬 버튼 찾기
                try:
                    new_review_button = driver.find_element(By.XPATH, "//a[contains(text(), '최신순')]")
                    self.update_log("✅ 최신순 버튼을 찾았습니다.")
                    driver.execute_script("arguments[0].click();", new_review_button)
                    self.update_log("✅ 최신순 정렬 버튼을 클릭했습니다.")
                    time.sleep(2)  # 정렬 변경 후 대기
                except:
                    self.update_log("🔄 대안 셀렉터로 최신순 버튼을 찾습니다...")
                    try:
                        new_review_button = driver.find_element(By.CSS_SELECTOR, 'a[data-shp-contents-id="최신순"]')
                        driver.execute_script("arguments[0].click();", new_review_button)
                        self.update_log("✅ 최신순 정렬 버튼을 클릭했습니다.")
                        time.sleep(2)  # 정렬 변경 후 대기
                    except Exception as e:
                        self.update_log(f"❌ 최신순 정렬 버튼을 찾을 수 없습니다: {e}")
            elif self.sort_option == "lowest":
                self.update_log("⭐ 평점 낮은순 정렬을 시도합니다...")
                # 평점 낮은 순 정렬 버튼 찾기
                try:
                    bad_review_button = driver.find_element(By.XPATH, "//a[contains(text(), '평점 낮은순')]")
                    self.update_log("✅ 평점 낮은순 버튼을 찾았습니다.")
                    driver.execute_script("arguments[0].click();", bad_review_button)
                    self.update_log("✅ 평점 낮은 순 정렬 버튼을 클릭했습니다.")
                    time.sleep(2)  # 정렬 변경 후 대기
                except:
                    self.update_log("🔄 대안 셀렉터로 평점 낮은순 버튼을 찾습니다...")
                    try:
                        bad_review_button = driver.find_element(By.CSS_SELECTOR, 'a[data-shp-contents-id="평점 낮은순"]')
                        driver.execute_script("arguments[0].click();", bad_review_button)
                        self.update_log("✅ 평점 낮은 순 정렬 버튼을 클릭했습니다.")
                        time.sleep(2)  # 정렬 변경 후 대기
                    except Exception as e:
                        self.update_log(f"❌ 평점 낮은 순 정렬 버튼을 찾을 수 없습니다: {e}")
            else:
                self.update_log(f"📊 기본 정렬 옵션 사용: {self.sort_option}")
        except Exception as e:
            self.update_log(f"❌ 정렬 옵션 처리 중 오류: {e}")

        time.sleep(2)

        self.update_log("📜 정렬 후 페이지 하단으로 스크롤합니다...")
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        self.update_log("✅ 페이지 하단으로 스크롤했습니다.")

        # 리뷰 추출 (리뷰 탭 클릭 후 새로운 셀렉터 사용)
        self.update_log("🔍 리뷰 목록의 셀렉터를 찾는 중...")
        
        # 리뷰 로딩을 위한 추가 대기
        self.update_log("⏳ 리뷰 목록 로딩을 위해 5초 추가 대기...")
        time.sleep(5)
        
        # 페이지 소스 일부를 확인하여 디버깅 정보 제공
        try:
            page_source = driver.page_source
            if "ul.RR2FSL9wTc" in page_source:
                self.update_log("✅ HTML에서 리뷰 목록 구조를 확인했습니다.")
            else:
                self.update_log("⚠️ HTML에서 리뷰 목록 구조를 찾을 수 없습니다.")
            
            if "li.PxsZltB5tV" in page_source:
                self.update_log("✅ HTML에서 리뷰 아이템 구조를 확인했습니다.")
            else:
                self.update_log("⚠️ HTML에서 리뷰 아이템 구조를 찾을 수 없습니다.")
                
            # 새로운 구조 확인
            if "ul.HTT4L8U0CU" in page_source:
                self.update_log("✅ HTML에서 새로운 리뷰 목록 구조를 확인했습니다.")
            else:
                self.update_log("⚠️ HTML에서 새로운 리뷰 목록 구조를 찾을 수 없습니다.")
                
        except Exception as e:
            self.update_log(f"⚠️ 페이지 소스 확인 실패: {e}")
        
        # 리뷰 목록의 셀렉터를 동적으로 찾기
        review_selector = self.find_review_selector(driver)
        if review_selector:
            self.update_log("✅ 리뷰 셀렉터를 찾았습니다. 리뷰 크롤링을 시작합니다.")
            self.crawl_reviews_until_page(driver, last_page_to_crawl, self.folder_path, review_selector)
        else:
            self.update_log("❌ 리뷰 셀렉터를 찾을 수 없습니다.")
            # 추가 디버깅 정보 제공
            self.update_log("🔍 디버깅을 위해 페이지 구조를 확인합니다...")
            try:
                # REVIEW 섹션의 모든 요소 확인
                review_section = driver.find_element(By.CSS_SELECTOR, "#REVIEW")
                self.update_log(f"✅ REVIEW 섹션 발견: {review_section.tag_name}")
                
                # 모든 ul 요소 확인
                ul_elements = driver.find_elements(By.CSS_SELECTOR, "#REVIEW ul")
                self.update_log(f"📋 REVIEW 섹션 내 ul 요소 수: {len(ul_elements)}")
                
                for i, ul in enumerate(ul_elements):
                    ul_class = ul.get_attribute('class')
                    li_count = len(ul.find_elements(By.CSS_SELECTOR, 'li'))
                    self.update_log(f"  ul {i+1}: class='{ul_class}', li 개수={li_count}")
                    
                    # 리뷰 목록인지 확인 (많은 li 요소를 가진 ul)
                    if li_count > 10:  # 리뷰 목록은 보통 10개 이상의 li를 가짐
                        self.update_log(f"    🎯 잠재적 리뷰 목록 발견: {ul_class}")
                        # 이 ul의 li들을 확인
                        li_elements_in_ul = ul.find_elements(By.CSS_SELECTOR, 'li')
                        for j, li in enumerate(li_elements_in_ul[:3]):  # 처음 3개만 확인
                            li_class = li.get_attribute('class')
                            li_html = li.get_attribute('outerHTML')[:100]
                            self.update_log(f"      li {j+1}: class='{li_class}' (HTML: {li_html}...)")
                
                # 모든 li 요소 확인
                li_elements = driver.find_elements(By.CSS_SELECTOR, "#REVIEW li")
                self.update_log(f"📋 REVIEW 섹션 내 li 요소 수: {len(li_elements)}")
                
                for i, li in enumerate(li_elements[:5]):  # 처음 5개만 확인
                    li_class = li.get_attribute('class')
                    self.update_log(f"  li {i+1}: class='{li_class}'")
                    
            except Exception as e:
                self.update_log(f"❌ 디버깅 정보 수집 실패: {e}")

        # 드라이버 종료
        self.update_log("🔚 Chrome 드라이버를 종료합니다...")
        driver.quit()
        self.update_log("✅ Chrome 드라이버 종료 완료")

        self.update_log("=" * 50)
        self.update_log("🎉 크롤링이 완료되었습니다!")
        self.update_log("=" * 50)

    def close_alert_if_present(self, driver):
        try:
            alert = driver.switch_to.alert
            alert.dismiss()
            self.update_log("예기치 않은 알림창을 닫았습니다.")
        except:
            pass

    def find_review_selector(self, driver):
        """리뷰 목록의 셀렉터를 동적으로 찾기 (새로운 구조)"""
        # 리뷰 섹션이 로드될 때까지 대기
        try:
            WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "#REVIEW"))
            )
            self.update_log("리뷰 섹션을 찾았습니다.")
        except:
            self.update_log("리뷰 섹션을 찾을 수 없습니다.")
        
        # HTML 구조를 더 정확히 분석하여 셀렉터 우선순위 조정
        possible_selectors = [
            # 가장 정확한 셀렉터들 (HTML 구조 기반)
            "ul.RR2FSL9wTc li.PxsZltB5tV",  # 메인 리뷰 목록
            "li.PxsZltB5tV",  # 리뷰 아이템들
            "#REVIEW ul.RR2FSL9wTc li",  # 리뷰 섹션 내의 정확한 리뷰 목록
            # 새로운 구조 시도 (로그에서 확인된 구조)
            "ul.HTT4L8U0CU li",  # 새로운 리뷰 목록 구조
            "#REVIEW ul.HTT4L8U0CU li",  # 리뷰 섹션 내의 새로운 리뷰 목록
            # 범용적인 셀렉터들
            "#REVIEW ul li",  # 리뷰 섹션 내의 li 요소들
            "#REVIEW li",  # 리뷰 섹션 내의 모든 li
            # 추가적인 셀렉터들
            "li[class*='PxsZltB5tV']",
            "ul[class*='RR2FSL9wTc'] li",
            "li[class*='PxsZltB5tV'][class*='_nlog_click']",
            ".review_list li",
            "[data-testid='review-item']",
            ".review-item",
            "ul li[class*='review']",
            "li[class*='review']",
            "div[class*='review'] li"
        ]
        
        for selector in possible_selectors:
            try:
                elements = driver.find_elements(By.CSS_SELECTOR, selector)
                self.update_log(f"셀렉터 '{selector}' 테스트: {len(elements)}개 요소 발견")
                if elements and len(elements) > 0:
                    # 실제 리뷰인지 확인 (리뷰 점수가 있는지 체크)
                    valid_reviews = 0
                    for i, elem in enumerate(elements[:3]):  # 처음 3개만 확인
                        try:
                            # 리뷰 점수 요소가 있는지 확인 (HTML 구조에 맞는 셀렉터 우선)
                            score_elem = None
                            score_selectors = [
                                'em.n6zq2yy0KA',  # HTML에서 확인된 정확한 클래스
                                'div.AlfkEF45qI em.n6zq2yy0KA',  # 더 구체적인 경로
                                'div.AlfkEF45qI em',
                                'em[class*="score"]',
                                'em[class*="star"]',
                                'em[class*="rating"]',
                                'em',
                                'span[class*="score"]',
                                'span[class*="star"]'
                            ]
                            
                            for score_selector in score_selectors:
                                try:
                                    score_elem = elem.find_element(By.CSS_SELECTOR, score_selector)
                                    if score_elem and score_elem.text and score_elem.text.strip().isdigit():
                                        break
                                except:
                                    continue
                            
                            if score_elem and score_elem.text and score_elem.text.strip().isdigit():
                                valid_reviews += 1
                                self.update_log(f"  요소 {i+1}: 점수 '{score_elem.text}' 발견")
                            else:
                                # 점수 요소를 찾지 못한 경우, 요소의 HTML 구조를 확인
                                try:
                                    elem_html = elem.get_attribute('outerHTML')[:200]  # 처음 200자만
                                    self.update_log(f"  요소 {i+1}: 점수 요소 없음 (HTML: {elem_html}...)")
                                except:
                                    self.update_log(f"  요소 {i+1}: 점수 요소 없음")
                        except Exception as e:
                            self.update_log(f"  요소 {i+1}: 점수 확인 실패 - {e}")
                            continue
                    
                    if valid_reviews > 0:
                        self.update_log(f"✅ 리뷰 셀렉터 발견: {selector} ({len(elements)}개 리뷰, {valid_reviews}개 유효)")
                        return selector
                    else:
                        self.update_log(f"❌ 셀렉터 '{selector}': 유효한 리뷰 없음")
            except Exception as e:
                self.update_log(f"❌ 셀렉터 '{selector}' 테스트 실패: {e}")
                continue
        
        # XPath로도 시도
        xpath_selectors = [
            "//ul[contains(@class, 'RR2FSL9wTc')]/li[contains(@class, 'PxsZltB5tV')]",
            "//div[@id='REVIEW']//ul[contains(@class, 'RR2FSL9wTc')]//li",
            # 새로운 구조 시도
            "//ul[contains(@class, 'HTT4L8U0CU')]//li",
            "//div[@id='REVIEW']//ul[contains(@class, 'HTT4L8U0CU')]//li",
            # 범용적인 XPath
            "//div[@id='REVIEW']//ul//li",
            "//li[contains(@class, 'PxsZltB5tV')]"
        ]
        
        for xpath in xpath_selectors:
            try:
                elements = driver.find_elements(By.XPATH, xpath)
                self.update_log(f"XPath '{xpath}' 테스트: {len(elements)}개 요소 발견")
                
                if elements and len(elements) > 0:
                    # 실제 리뷰인지 확인
                    valid_reviews = 0
                    for i, elem in enumerate(elements[:3]):
                        try:
                            # 리뷰 점수 요소가 있는지 확인 (HTML 구조에 맞는 셀렉터 우선)
                            score_elem = None
                            score_selectors = [
                                'em.n6zq2yy0KA',  # HTML에서 확인된 정확한 클래스
                                'div.AlfkEF45qI em.n6zq2yy0KA',  # 더 구체적인 경로
                                'div.AlfkEF45qI em',
                                'em[class*="score"]',
                                'em[class*="star"]',
                                'em[class*="rating"]',
                                'em',
                                'span[class*="score"]',
                                'span[class*="star"]'
                            ]
                            
                            for score_selector in score_selectors:
                                try:
                                    score_elem = elem.find_element(By.CSS_SELECTOR, score_selector)
                                    if score_elem and score_elem.text and score_elem.text.strip().isdigit():
                                        break
                                except:
                                    continue
                            
                            if score_elem and score_elem.text and score_elem.text.strip().isdigit():
                                valid_reviews += 1
                                self.update_log(f"  요소 {i+1}: 점수 '{score_elem.text}' 발견")
                            else:
                                # 점수 요소를 찾지 못한 경우, 요소의 HTML 구조를 확인
                                try:
                                    elem_html = elem.get_attribute('outerHTML')[:200]  # 처음 200자만
                                    self.update_log(f"  요소 {i+1}: 점수 요소 없음 (HTML: {elem_html}...)")
                                except:
                                    self.update_log(f"  요소 {i+1}: 점수 요소 없음")
                        except Exception as e:
                            self.update_log(f"  요소 {i+1}: 점수 확인 실패 - {e}")
                            continue
                    
                    if valid_reviews > 0:
                        self.update_log(f"✅ 리뷰 셀렉터 발견 (XPath): {xpath} ({len(elements)}개 리뷰, {valid_reviews}개 유효)")
                        return xpath
                    else:
                        self.update_log(f"❌ XPath '{xpath}': 유효한 리뷰 없음")
            except Exception as e:
                self.update_log(f"❌ XPath '{xpath}' 테스트 실패: {e}")
                continue
            
        # 마지막 시도: 리뷰가 로드될 때까지 대기하고 다시 시도
        self.update_log("🔄 리뷰 로딩을 위해 추가 대기 후 재시도...")
        time.sleep(10)  # 10초 추가 대기
        
        # 다시 시도
        for selector in possible_selectors[:5]:  # 상위 5개 셀렉터만 재시도
            try:
                elements = driver.find_elements(By.CSS_SELECTOR, selector)
                self.update_log(f"재시도 - 셀렉터 '{selector}' 테스트: {len(elements)}개 요소 발견")
                if elements and len(elements) > 0:
                    # 실제 리뷰인지 확인
                    valid_reviews = 0
                    for i, elem in enumerate(elements[:3]):
                        try:
                            score_elem = None
                            score_selectors = [
                                'em.n6zq2yy0KA',
                                'div.AlfkEF45qI em.n6zq2yy0KA',
                                'div.AlfkEF45qI em',
                                'em[class*="score"]',
                                'em[class*="star"]',
                                'em[class*="rating"]',
                                'em',
                                'span[class*="score"]',
                                'span[class*="star"]'
                            ]
                            
                            for score_selector in score_selectors:
                                try:
                                    score_elem = elem.find_element(By.CSS_SELECTOR, score_selector)
                                    if score_elem and score_elem.text and score_elem.text.strip().isdigit():
                                        break
                                except:
                                    continue
                            
                            if score_elem and score_elem.text and score_elem.text.strip().isdigit():
                                valid_reviews += 1
                                self.update_log(f"  재시도 - 요소 {i+1}: 점수 '{score_elem.text}' 발견")
                            else:
                                self.update_log(f"  재시도 - 요소 {i+1}: 점수 요소 없음")
                        except Exception as e:
                            self.update_log(f"  재시도 - 요소 {i+1}: 점수 확인 실패 - {e}")
                            continue
                    
                    if valid_reviews > 0:
                        self.update_log(f"✅ 재시도 성공 - 리뷰 셀렉터 발견: {selector} ({len(elements)}개 리뷰, {valid_reviews}개 유효)")
                        return selector
            except Exception as e:
                self.update_log(f"❌ 재시도 - 셀렉터 '{selector}' 테스트 실패: {e}")
                continue
        
        self.update_log("리뷰 셀렉터를 찾을 수 없습니다.")
        return None

    def crawl_reviews_until_page(self, driver, last_page, photo_folder_path, review_selector):
        self.update_log("=" * 50)
        self.update_log("📚 리뷰 페이지 크롤링을 시작합니다...")
        self.update_log("=" * 50)
        current_page = 1
        chunk_count = 1
        while True:
            self.update_log(f"📄 페이지 {current_page} 크롤링 시작...")
            chunk_reviews = []
            chunk_start_page = current_page
            
            for _ in range(self.chunk_size):
                if (last_page != "max" and current_page > last_page):
                    self.update_log(f"📊 설정된 페이지 수({last_page})에 도달했습니다.")
                    break
                
                self.update_log(f"🔍 페이지 {current_page}에서 리뷰를 추출합니다...")
                reviews_data = self.extract_reviews(driver, photo_folder_path, current_page, review_selector)
                chunk_reviews.extend(reviews_data)
                self.update_log(f"📊 페이지 {current_page}에서 {len(reviews_data)}개 리뷰 추출 완료")
                
                if not reviews_data:  # 더 이상 리뷰가 없으면 종료
                    self.update_log("⚠️ 더 이상 리뷰가 없습니다. 크롤링을 종료합니다.")
                    break
                
                current_page += 1
                
                try:
                    self.update_log(f"🔍 페이지 {current_page}로 이동을 시도합니다...")
                    # 페이지네이션 버튼들 찾기
                    pagination_selectors = [
                        'a.hyY6CXtbcn[aria-current="true"]',  # 현재 페이지
                        'a.hyY6CXtbcn',  # 페이지 번호들
                        'a.JY2WGJ4hXh.I3i1NSoFdB',  # 다음 버튼
                        'a[aria-label="다음 페이지"]',  # 다음 버튼 대안
                        'a[title="다음"]'  # 다음 버튼 대안
                    ]
                    
                    # 현재 페이지 확인
                    self.update_log(f"  📍 현재 페이지를 확인합니다...")
                    current_page_element = None
                    for selector in pagination_selectors[:1]:  # 현재 페이지만 확인
                        try:
                            current_page_element = driver.find_element(By.CSS_SELECTOR, selector)
                            break
                        except:
                            continue
                    
                    if current_page_element:
                        current_page_number = int(current_page_element.text)
                        self.update_log(f"  ✅ 현재 페이지: {current_page_number}")
                    else:
                        self.update_log(f"  ⚠️ 현재 페이지를 확인할 수 없습니다.")
                    
                    # 다음 페이지로 이동 시도
                    self.update_log(f"  🔄 다음 페이지로 이동을 시도합니다...")
                    next_page_clicked = False
                    
                    # 1. 다음 페이지 번호 버튼 클릭 시도
                    self.update_log(f"    📄 페이지 번호 버튼을 찾는 중...")
                    try:
                        next_page_elements = driver.find_elements(By.CSS_SELECTOR, 'a.hyY6CXtbcn[aria-current="false"]')
                        for elem in next_page_elements:
                            if elem.text.isdigit() and int(elem.text) == current_page:
                                driver.execute_script("arguments[0].click();", elem)
                                next_page_clicked = True
                                self.update_log(f"    ✅ 페이지 {current_page}로 이동했습니다.")
                                break
                    except Exception as e:
                        self.update_log(f"    ❌ 페이지 번호 클릭 실패: {e}")
                    
                    # 2. 다음 버튼 클릭 시도
                    if not next_page_clicked:
                        self.update_log(f"    🔄 다음 버튼을 찾는 중...")
                        for selector in pagination_selectors[2:]:  # 다음 버튼들만 확인
                            try:
                                next_button = driver.find_element(By.CSS_SELECTOR, selector)
                                driver.execute_script("arguments[0].click();", next_button)
                                next_page_clicked = True
                                self.update_log("    ✅ 다음 버튼을 클릭했습니다.")
                                break
                            except:
                                continue
                    
                    if not next_page_clicked:
                        self.update_log("    ❌ 다음 페이지로 이동할 수 없습니다.")
                        break
                    
                    self.update_log(f"    ⏳ 페이지 로딩을 위해 {random.uniform(2, 4):.1f}초 대기...")
                    time.sleep(random.uniform(2, 4))  # 페이지 로딩 대기
                    
                except Exception as e:
                    self.update_log(f"❌ 페이지 이동 실패: {e}")
                    break
            
            # 청크 데이터 저장
            if chunk_reviews:
                self.update_log(f"💾 청크 {chunk_count} 데이터를 저장합니다...")
                excel_path = os.path.join(self.folder_path, f'reviews_chunk_{chunk_count}.xlsx')
                self.save_to_excel(chunk_reviews, excel_path, chunk_start_page)
                self.update_log(f"✅ 청크 {chunk_count} 저장 완료: {len(chunk_reviews)}개 리뷰")
                chunk_count += 1
            else:
                self.update_log("⚠️ 저장할 리뷰 데이터가 없습니다.")
            
            if (last_page != "max" and current_page > last_page) or not reviews_data:
                self.update_log("🏁 크롤링 조건을 만족하여 종료합니다.")
                break

    def extract_reviews(self, driver, photo_folder_path, current_page, review_selector):
        self.update_log(f"📄 페이지 {current_page}에서 리뷰를 추출합니다...")
        reviews_data = []
        try:
            # CSS 셀렉터인지 XPath인지 확인
            if review_selector.startswith("//"):
                self.update_log(f"🔍 XPath로 리뷰를 찾는 중: {review_selector}")
                reviews = driver.find_elements(By.XPATH, review_selector)
            else:
                self.update_log(f"🔍 CSS 셀렉터로 리뷰를 찾는 중: {review_selector}")
                reviews = driver.find_elements(By.CSS_SELECTOR, review_selector)
            
            self.update_log(f"✅ 리뷰 {len(reviews)}개를 찾았습니다.")
            
            if not reviews:
                self.update_log("리뷰를 찾을 수 없습니다. 페이지 구조가 변경되었을 수 있습니다.")
                return reviews_data
            
            for index, review in enumerate(reviews):
                self.update_log(f"📝 리뷰 {index+1}/{len(reviews)} 추출 중...")
                try:
                    # 리뷰 점수 추출 (HTML 구조에 맞는 셀렉터 우선)
                    self.update_log(f"  ⭐ 리뷰 점수 추출 중...")
                    review_score = self.safe_extract_text(review, [
                        'em.n6zq2yy0KA',  # HTML에서 확인된 정확한 클래스
                        'div.AlfkEF45qI em.n6zq2yy0KA',  # 더 구체적인 경로
                        'div.AlfkEF45qI em',
                        'em[class*="score"]',
                        'em[class*="star"]',
                        'em[class*="rating"]',
                        'em',
                        '.score',
                        'span[class*="score"]',
                        'span[class*="star"]'
                    ], "점수 없음")
                    self.update_log(f"  ✅ 리뷰 점수: {review_score}")
                    
                    # 리뷰어 이름 추출 (HTML 구조에 맞는 셀렉터 우선)
                    self.update_log(f"  👤 리뷰어 이름 추출 중...")
                    reviewer_name = self.safe_extract_text(review, [
                        'strong.MX91DFZo2F',  # HTML에서 확인된 정확한 클래스
                        'div.AlfkEF45qI strong.MX91DFZo2F',  # 더 구체적인 경로
                        'div.Db9Dtnf7gY strong.MX91DFZo2F',
                        'strong[class*="name"]',
                        '.reviewer-name',
                        'span[class*="name"]'
                    ], "이름 없음")
                    self.update_log(f"  ✅ 리뷰어: {reviewer_name}")
                    
                    # 리뷰 날짜 추출 (HTML 구조에 맞는 셀렉터 우선)
                    self.update_log(f"  📅 리뷰 날짜 추출 중...")
                    review_date = self.safe_extract_text(review, [
                        'span.MX91DFZo2F',  # HTML에서 확인된 정확한 클래스
                        'div.Db9Dtnf7gY span.MX91DFZo2F',  # 더 구체적인 경로
                        'div.AlfkEF45qI div.Db9Dtnf7gY span.MX91DFZo2F',
                        'span[class*="date"]',
                        '.review-date',
                        'div[class*="date"]'
                    ], "날짜 없음")
                    self.update_log(f"  ✅ 리뷰 날짜: {review_date}")
                    
                    # 상품 옵션 정보 추출 (HTML 구조에 맞는 셀렉터 우선)
                    self.update_log(f"  🛍️ 상품 옵션 정보 추출 중...")
                    product_name = self.safe_extract_text(review, [
                        'div.b_caIle8kC',  # HTML에서 확인된 정확한 클래스
                        'div.AlfkEF45qI div.b_caIle8kC',  # 더 구체적인 경로
                        'div[class*="product"]',
                        '.product-name',
                        'span[class*="product"]'
                    ], "정보 없음")
                    self.update_log(f"  ✅ 상품 옵션: {product_name}")

                    # 리뷰 내용 추출 (여러 방법 시도)
                    self.update_log(f"  📝 리뷰 내용 추출 중...")
                    content = self.safe_extract_review_content(review)
                    self.update_log(f"  ✅ 리뷰 내용 길이: {len(content)}자")

                    # 리뷰 유형 확인
                    self.update_log(f"  🏷️ 리뷰 유형 확인 중...")
                    review_type = self.determine_review_type(review)
                    self.update_log(f"  ✅ 리뷰 유형: {review_type}")

                    # 리뷰 이미지 추출
                    self.update_log(f"  📸 리뷰 이미지 추출 중...")
                    photos = self.extract_review_photos(review, photo_folder_path, current_page, index)
                    self.update_log(f"  ✅ 추출된 이미지 수: {len(photos)}개")

                    reviews_data.append({
                        "Review Score": review_score,
                        "Reviewer Name": reviewer_name,
                        "Review Date": review_date,
                        "Product(Option) Name": product_name,
                        "Review Type": review_type,
                        "Content": content,
                        "Photos": photos
                    })
                    self.update_log(f"  ✅ 리뷰 추출 완료: 페이지 {current_page}, 리뷰 {index+1}")
                    
                except Exception as e:
                    self.update_log(f"  ❌ 리뷰 {index+1} 추출 실패: {e}")
                    continue
                    
        except Exception as e:
            self.update_log(f"리뷰 정보 추출 실패: {e}")
        return reviews_data

    def safe_extract_text(self, element, selectors, default_value=""):
        """여러 셀렉터를 시도하여 텍스트를 안전하게 추출"""
        for selector in selectors:
            try:
                found_element = element.find_element(By.CSS_SELECTOR, selector)
                text = found_element.text.strip()
                if text:
                    return text
            except:
                continue
        return default_value

    def safe_extract_review_content(self, review):
        """리뷰 내용을 안전하게 추출 (HTML 구조에 맞는 셀렉터 우선)"""
        content_selectors = [
            'div.KqJ8Qqw082 span.MX91DFZo2F',  # HTML에서 확인된 정확한 경로
            'div.AlfkEF45qI div.KqJ8Qqw082 span.MX91DFZo2F',  # 더 구체적인 경로
            'span.MX91DFZo2F',
            'div[class*="content"] span',
            '.review-content',
            'p[class*="content"]',
            'div[class*="text"]'
        ]
        
        for selector in content_selectors:
            try:
                content_spans = review.find_elements(By.CSS_SELECTOR, selector)
                if content_spans:
                    # 가장 긴 텍스트를 리뷰 내용으로 선택
                    contents = [span.text.strip() for span in content_spans if span.text.strip()]
                    if contents:
                        # 리뷰 내용은 보통 가장 긴 텍스트이므로 길이로 판단
                        longest_content = max(contents, key=len)
                        # 너무 짧은 텍스트는 제외 (날짜나 이름 등)
                        if len(longest_content) > 10:
                            return longest_content
            except:
                continue
        return ""

    def determine_review_type(self, review):
        """리뷰 유형을 결정 (HTML 구조에 맞는 셀렉터 우선)"""
        review_type = "일반리뷰"
        try:
            type_selectors = [
                'span.W1IZsaUmnu',  # HTML에서 확인된 정확한 클래스
                'div.KqJ8Qqw082 span.W1IZsaUmnu',  # 더 구체적인 경로
                'div.AlfkEF45qI div.KqJ8Qqw082 span.W1IZsaUmnu',
                'div.KqJ8Qqw082 span',
                'div.AlfkEF45qI div.KqJ8Qqw082 span',
                'span[class*="tag"]',
                '.review-tag',
                'div[class*="type"]'
            ]
            
            for selector in type_selectors:
                try:
                    type_elements = review.find_elements(By.CSS_SELECTOR, selector)
                    types = [elem.text.strip() for elem in type_elements if elem.text.strip()]
                    
                    # 한달사용, 재구매 등 확인
                    is_one_month_review = any("한달사용" in t for t in types)
                    is_reorder = any("재구매" in t for t in types)
                    
                    if is_one_month_review and is_reorder:
                        review_type = "한달+재구매"
                    elif is_one_month_review:
                        review_type = "한달사용기"
                    elif is_reorder:
                        review_type = "재구매"
                    break
                except:
                    continue
        except Exception as e:
            self.update_log(f"리뷰 유형 확인 실패: {e}")
        return review_type

    def extract_review_photos(self, review, photo_folder_path, current_page, review_index):
        """리뷰 이미지를 안전하게 추출"""
        self.update_log(f"      📸 리뷰 이미지 추출 시작...")
        photos = []
        try:
            photo_selectors = [
                'img.UpImHAUeYJ[alt="review_image"]',  # HTML에서 확인된 정확한 클래스
                'div.AlfkEF45qI img.UpImHAUeYJ[alt="review_image"]',  # 더 구체적인 경로
                'div.s30AvhHfb0 img.UpImHAUeYJ[alt="review_image"]',  # 이미지 컨테이너 내부
                'img[alt="review_image"]',
                'img[class*="review"]',
                '.review-photo img'
            ]
            
            self.update_log(f"      🔍 이미지 요소를 찾는 중...")
            photo_elements = []
            for selector in photo_selectors:
                try:
                    elements = review.find_elements(By.CSS_SELECTOR, selector)
                    if elements:
                        photo_elements = elements
                        self.update_log(f"      ✅ 이미지 요소 {len(elements)}개 발견: {selector}")
                        break
                except:
                    continue
            
            if not photo_elements:
                self.update_log(f"      ⚠️ 이미지 요소를 찾을 수 없습니다.")
                return photos
            
            for i, photo_element in enumerate(photo_elements):
                self.update_log(f"        🖼️ 이미지 {i+1}/{len(photo_elements)} 처리 중...")
                try:
                    # data-src 속성을 우선적으로 사용 (원본 이미지)
                    self.update_log(f"          🔗 이미지 URL 추출 중...")
                    photo_url = photo_element.get_attribute('data-src')
                    if not photo_url:
                        photo_url = photo_element.get_attribute('src')
                    
                    if not photo_url:
                        self.update_log(f"          ⚠️ 이미지 URL을 찾을 수 없습니다.")
                        continue
                    
                    self.update_log(f"          📝 원본 URL: {photo_url}")
                    
                    # 네이버 이미지 URL에서 파라미터 제거하여 원본 가져오기
                    if 'pstatic.net' in photo_url:
                        original_url = photo_url.split('?')[0]
                    else:
                        original_url = photo_url
                    
                    self.update_log(f"          🔄 처리된 URL: {original_url}")
                    
                    # 이미지 다운로드
                    self.update_log(f"          ⬇️ 이미지 다운로드 중...")
                    response = requests.get(original_url, timeout=10)
                    response.raise_for_status()
                    
                    photo_path = os.path.join(photo_folder_path, f'review_page{current_page}_{review_index+1}_photo_{i+1}.jpg')
                    with open(photo_path, 'wb') as file:
                        file.write(response.content)
                    
                    self.update_log(f"          💾 이미지 저장 완료: {photo_path}")
                    
                    # 이미지 크기 확인
                    try:
                        with PILImage.open(photo_path) as img:
                            width, height = img.size
                            self.update_log(f"          ✅ 이미지 저장 완료: {width}x{height}")
                    except Exception as e:
                        self.update_log(f"          ⚠️ 이미지 크기 확인 실패: {e}")
                    
                    photos.append(photo_path)
                    self.update_log(f"          ✅ 이미지 {i+1} 처리 완료")
                    
                except Exception as e:
                    self.update_log(f"          ❌ 이미지 {i+1} 다운로드 실패: {e}")
                    continue
                    
        except Exception as e:
            self.update_log(f"      ❌ 이미지 추출 실패: {e}")
        
        self.update_log(f"      📊 총 {len(photos)}개 이미지 추출 완료")
        return photos

    def save_to_excel(self, reviews_data, excel_path, start_page):
        # 기존 데이터 읽기 (중복 체크용)
        existing_reviews = set()
        if os.path.exists(excel_path):
            wb = load_workbook(excel_path)
            ws = wb.active
            start_row = ws.max_row + 1
            
            # 기존 리뷰 데이터를 set으로 저장 (중복 체크용)
            for row in range(2, ws.max_row + 1):  # 헤더 제외
                reviewer = ws.cell(row=row, column=3).value or ""
                date = ws.cell(row=row, column=4).value or ""
                score = ws.cell(row=row, column=2).value or ""
                content = ws.cell(row=row, column=7).value or ""
                # 중복 체크용 키 생성 (내용 전체)
                review_key = f"{reviewer}_{date}_{score}_{content}"
                existing_reviews.add(review_key)
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Reviews"
            headers = ["Page", "Review Score", "Reviewer Name", "Review Date", "Product(Option) Name", "Review Type", "Content", "Photo"]
            ws.append(headers)
            for i, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=i, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            start_row = 2

        # 컬럼 너비 설정
        column_widths = {
            "A": 10, "B": 12, "C": 15, "D": 15, "E": 25, "F": 15, "G": 60, "H": 12
        }
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        # 중복되지 않은 리뷰만 필터링
        new_reviews = []
        duplicate_count = 0
        
        for review in reviews_data:
            # 중복 체크용 키 생성 (내용 전체)
            reviewer = review["Reviewer Name"] or ""
            date = review["Review Date"] or ""
            score = review["Review Score"] or ""
            content = review["Content"] or ""
            review_key = f"{reviewer}_{date}_{score}_{content}"
            
            if review_key not in existing_reviews:
                new_reviews.append(review)
                existing_reviews.add(review_key)  # 다음 중복 체크를 위해 추가
            else:
                duplicate_count += 1
        
        if duplicate_count > 0:
            self.update_log(f"중복 리뷰 {duplicate_count}개를 제외했습니다.")

        # 리뷰 데이터 추가
        for row_index, review in enumerate(new_reviews, start=start_row):
            ws.cell(row=row_index, column=1, value=start_page).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=row_index, column=2, value=review["Review Score"]).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.cell(row=row_index, column=3, value=review["Reviewer Name"]).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.cell(row=row_index, column=4, value=review["Review Date"]).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.cell(row=row_index, column=5, value=review["Product(Option) Name"]).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.cell(row=row_index, column=6, value=review["Review Type"]).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.cell(row=row_index, column=7, value=review["Content"]).alignment = Alignment(vertical="center", wrap_text=True)

            start_page += 1

        # 행 높이 설정
        for row in range(start_row, ws.max_row + 1):
            ws.row_dimensions[row].height = 75

        wb.save(excel_path)
        self.update_log(f"새로운 리뷰 {len(new_reviews)}개를 {excel_path}에 저장했습니다.")

class ReviewCrawlerGUI:
    def __init__(self, root):
        self.root = root
        self.root.title('리뷰 수집기')
        self.root.geometry('500x600')
        
        # 변수들
        self.site_var = tk.StringVar(value="naver")
        self.sort_var = tk.StringVar(value="ranking")
        self.page_var = tk.StringVar(value="5")
        self.custom_page_var = tk.StringVar()
        self.open_folder_var = tk.BooleanVar()
        
        self.initUI()

    def initUI(self):
        # 메인 프레임
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # 상품 링크
        ttk.Label(main_frame, text="상품 링크:").grid(row=0, column=0, sticky=tk.W, pady=5)
        self.url_input = ttk.Entry(main_frame, width=50)
        self.url_input.grid(row=0, column=1, columnspan=3, sticky=(tk.W, tk.E), pady=5)
        
        # 사이트 선택
        ttk.Label(main_frame, text="사이트:").grid(row=1, column=0, sticky=tk.W, pady=5)
        ttk.Radiobutton(main_frame, text="네이버", variable=self.site_var, value="naver").grid(row=1, column=1, pady=5)
        ttk.Radiobutton(main_frame, text="쿠팡", variable=self.site_var, value="coupang").grid(row=1, column=2, pady=5)
        
        # 정렬 순서
        ttk.Label(main_frame, text="정렬 순서:").grid(row=2, column=0, sticky=tk.W, pady=5)
        ttk.Radiobutton(main_frame, text="랭킹순", variable=self.sort_var, value="ranking").grid(row=2, column=1, pady=5)
        ttk.Radiobutton(main_frame, text="최신순", variable=self.sort_var, value="latest").grid(row=2, column=2, pady=5)
        ttk.Radiobutton(main_frame, text="평점낮은순", variable=self.sort_var, value="lowest").grid(row=2, column=3, pady=5)
        
        # 리뷰 페이지
        ttk.Label(main_frame, text="리뷰 페이지:").grid(row=3, column=0, sticky=tk.W, pady=5)
        ttk.Radiobutton(main_frame, text="5", variable=self.page_var, value="5").grid(row=3, column=1, pady=5)
        ttk.Radiobutton(main_frame, text="15", variable=self.page_var, value="15").grid(row=3, column=2, pady=5)
        ttk.Radiobutton(main_frame, text="50", variable=self.page_var, value="50").grid(row=3, column=3, pady=5)
        ttk.Radiobutton(main_frame, text="max", variable=self.page_var, value="max").grid(row=3, column=4, pady=5)
        
        # 직접 입력
        ttk.Label(main_frame, text="직접 입력:").grid(row=4, column=0, sticky=tk.W, pady=5)
        self.custom_page_entry = ttk.Entry(main_frame, textvariable=self.custom_page_var, width=10)
        self.custom_page_entry.grid(row=4, column=1, sticky=tk.W, pady=5)
        self.custom_page_entry.bind('<KeyRelease>', self.on_custom_page_change)
        
        # 저장 경로
        ttk.Label(main_frame, text="저장 경로:").grid(row=5, column=0, sticky=tk.W, pady=5)
        self.path_input = ttk.Entry(main_frame, width=40)
        self.path_input.grid(row=5, column=1, columnspan=2, sticky=(tk.W, tk.E), pady=5)
        ttk.Button(main_frame, text="선택", command=self.select_path).grid(row=5, column=3, pady=5)
        
        # 폴더 자동 열기 옵션
        ttk.Checkbutton(main_frame, text="크롤링 후 폴더 열기", variable=self.open_folder_var).grid(row=6, column=0, columnspan=4, sticky=tk.W, pady=5)
        
        # 실행 버튼
        self.run_button = ttk.Button(main_frame, text="실행", command=self.run_crawler)
        self.run_button.grid(row=7, column=0, columnspan=4, pady=10)
        
        # 로그 출력
        ttk.Label(main_frame, text="로그:").grid(row=8, column=0, sticky=tk.W, pady=5)
        self.log_text = scrolledtext.ScrolledText(main_frame, height=15, width=60)
        self.log_text.grid(row=9, column=0, columnspan=4, sticky=(tk.W, tk.E, tk.N, tk.S), pady=5)
        
        # 그리드 가중치 설정
        main_frame.columnconfigure(1, weight=1)
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)

    def on_custom_page_change(self, event):
        if self.custom_page_var.get():
            self.page_var.set("custom")

    def select_path(self):
        folder_path = filedialog.askdirectory(title="저장 경로 선택")
        if folder_path:
            self.path_input.delete(0, tk.END)
            self.path_input.insert(0, folder_path)

    def update_log(self, message):
        self.log_text.insert(tk.END, message + "\n")
        self.log_text.see(tk.END)
        self.root.update_idletasks()

    def run_crawler(self):
        # 입력값 가져오기
        input_url = self.url_input.get()
        if not input_url:
            messagebox.showerror("오류", "상품 링크를 입력해주세요.")
            return
            
        sort_option = self.sort_var.get()
        
        if self.page_var.get() == "custom":
            input_pages = self.custom_page_var.get()
            if not input_pages:
                messagebox.showerror("오류", "페이지 수를 입력해주세요.")
                return
        else:
            input_pages = self.page_var.get()

        photo_folder_path = self.path_input.get()
        if not photo_folder_path:
            messagebox.showerror("오류", "저장 경로를 선택해주세요.")
            return

        # 크롤링 스레드 시작
        self.crawler_thread = CrawlerThread(input_url, input_pages, photo_folder_path, sort_option, callback=self.update_log)
        self.crawler_thread.daemon = True
        self.crawler_thread.start()

        # UI 비활성화
        self.run_button.config(state='disabled')
        self.update_log("크롤링을 시작합니다...")
        
        # 완료 체크를 위한 타이머
        self.check_thread_completion()

    def check_thread_completion(self):
        if hasattr(self, 'crawler_thread') and self.crawler_thread.is_alive():
            self.root.after(1000, self.check_thread_completion)
        else:
            self.run_button.config(state='normal')
            self.update_log("크롤링이 완료되었습니다.")
            
            # 폴더 자동 열기
            if self.open_folder_var.get():
                folder_path = self.path_input.get()
                if os.path.exists(folder_path):
                    if sys.platform == 'win32':
                        os.startfile(folder_path)
                    elif sys.platform == 'darwin':  # macOS
                        subprocess.Popen(['open', folder_path])
                    else:  # linux
                        subprocess.Popen(['xdg-open', folder_path])

if __name__ == '__main__':
    root = tk.Tk()
    app = ReviewCrawlerGUI(root)
    root.mainloop()
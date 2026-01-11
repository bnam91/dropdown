import sys
import tkinter as tk
from tkinter import ttk, filedialog, scrolledtext, messagebox
import threading
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import random
import os
import requests
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment
import subprocess
from PIL import Image as PILImage
import io

class CrawlerThread(threading.Thread):
    def __init__(self, search_keyword, pages, folder_path, sort_option, chunk_size=50, callback=None):
        threading.Thread.__init__(self)
        self.search_keyword = search_keyword
        self.pages = pages
        self.folder_path = folder_path
        self.sort_option = sort_option
        self.chunk_size = chunk_size
        self.callback = callback

    def run(self):
        self.crawl_reviews()

    def update_log(self, message):
        if self.callback:
            self.callback(message)

    def crawl_reviews(self):
        self.update_log("=" * 50)
        self.update_log("ğŸš€ í¬ë¡¤ë§ì„ ì‹œì‘í•©ë‹ˆë‹¤...")
        self.update_log("=" * 50)

        # Selenium ì„¤ì •
        self.update_log("ğŸ“‹ Selenium ì˜µì…˜ì„ ì„¤ì •í•©ë‹ˆë‹¤...")
        options = Options()
        options.add_experimental_option("detach", True)
        options.add_argument("disable-blink-features=AutomationControlled")
        options.add_experimental_option("excludeSwitches", ["enable-logging"])
        self.update_log("âœ… Selenium ì˜µì…˜ ì„¤ì • ì™„ë£Œ")

        self.update_log("ğŸŒ Chrome ë“œë¼ì´ë²„ë¥¼ ì‹œì‘í•©ë‹ˆë‹¤...")
        driver = webdriver.Chrome(options=options)
        self.update_log("âœ… Chrome ë“œë¼ì´ë²„ ì‹œì‘ ì™„ë£Œ")

        # ë¨¼ì € ë„¤ì´ë²„ë¡œ ì´ë™
        self.update_log("ğŸ  ë„¤ì´ë²„ ë©”ì¸ í˜ì´ì§€ë¡œ ì´ë™í•©ë‹ˆë‹¤...")
        driver.get("https://www.naver.com")
        self.update_log("âœ… ë„¤ì´ë²„ì— ì ‘ì†í–ˆìŠµë‹ˆë‹¤.")
        time.sleep(random.uniform(2, 5))

        # ë³€ìˆ˜ ì •ì˜
        self.update_log("âš™ï¸ í¬ë¡¤ë§ ì„¤ì •ì„ ì´ˆê¸°í™”í•©ë‹ˆë‹¤...")
        search_keyword = self.search_keyword
        last_page_to_crawl = int(self.pages) if self.pages.isdigit() else self.pages
        sleep_time = 2
        self.update_log(f"ğŸ“Š ì„¤ì •ëœ í˜ì´ì§€ ìˆ˜: {last_page_to_crawl}")
        self.update_log(f"ğŸ” ê²€ìƒ‰ì–´: {search_keyword}")
        self.update_log(f"â±ï¸ ëŒ€ê¸° ì‹œê°„: {sleep_time}ì´ˆ")

        # ë„¤ì´ë²„ (ìƒˆë¡œìš´ êµ¬ì¡°)
        product_review_button_selector = 'a[data-name="REVIEW"]'
        new_review_button_selector = 'a[data-name="REVIEW"]'  # ë¦¬ë·° íƒ­ í´ë¦­ í›„ ì •ë ¬ ì˜µì…˜ í™•ì¸ í•„ìš”
        bad_review_button_selector = 'a[data-name="REVIEW"]'  # ë¦¬ë·° íƒ­ í´ë¦­ í›„ ì •ë ¬ ì˜µì…˜ í™•ì¸ í•„ìš”
        review_selector = "#REVIEW > div > div._2LvIMaBiIO > div._2g7PKvqCKe > ul > li"  # ë¦¬ë·° íƒ­ í´ë¦­ í›„ í™•ì¸ í•„ìš”

        # ë„¤ì´ë²„ ì ‘ì† í›„ 5ì´ˆ ëŒ€ê¸°
        self.update_log("â³ ë„¤ì´ë²„ ì ‘ì† í›„ 5ì´ˆ ëŒ€ê¸°í•©ë‹ˆë‹¤...")
        for i in range(5, 0, -1):
            self.update_log(f"â° ëŒ€ê¸° ì¤‘: {i}ì´ˆ")
            time.sleep(1)
        self.update_log("âœ… ë„¤ì´ë²„ ëŒ€ê¸° ì™„ë£Œ!")

        # ê²€ìƒ‰ í˜ì´ì§€ë¡œ ì´ë™ (ê²€ìƒ‰ì–´ë¥¼ URL ì¸ì½”ë”©í•˜ì—¬ ë™ì  ìƒì„±)
        import urllib.parse
        encoded_keyword = urllib.parse.quote(search_keyword)
        search_url = f"https://search.naver.com/search.naver?where=nexearch&sm=top_hty&fbm=0&ie=utf8&query={encoded_keyword}&ackey=ayy89dsf"
        self.update_log("ğŸ” ë„¤ì´ë²„ ê²€ìƒ‰ í˜ì´ì§€ë¡œ ì´ë™í•©ë‹ˆë‹¤...")
        self.update_log(f"ğŸ“ ê²€ìƒ‰ URL: {search_url}")
        driver.get(search_url)
        self.update_log("âœ… ê²€ìƒ‰ í˜ì´ì§€ì— ì ‘ì†í–ˆìŠµë‹ˆë‹¤.")
        
        # ì‚¬ìš©ìê°€ ì§ì ‘ ìƒí’ˆ í˜ì´ì§€ë¡œ ì´ë™í•  ì‹œê°„ ì œê³µ
        self.update_log("â³ ì‚¬ìš©ìê°€ ì§ì ‘ ìƒí’ˆ í˜ì´ì§€ë¡œ ì´ë™í•  ì‹œê°„ì„ ì œê³µí•©ë‹ˆë‹¤...")
        self.update_log("ğŸ“ ì´ ì‹œê°„ ë™ì•ˆ ë¸Œë¼ìš°ì €ì—ì„œ ì›í•˜ëŠ” ìƒí’ˆ í˜ì´ì§€ë¡œ ì´ë™í•´ì£¼ì„¸ìš”.")
        self.update_log("â° 15ì´ˆ í›„ ìë™ìœ¼ë¡œ í¬ë¡¤ë§ì„ ì‹œì‘í•©ë‹ˆë‹¤...")
        for i in range(15, 0, -1):
            self.update_log(f"â° ë‚¨ì€ ì‹œê°„: {i}ì´ˆ")
            time.sleep(1)
        self.update_log("âœ… ëŒ€ê¸° ì™„ë£Œ! í¬ë¡¤ë§ì„ ì‹œì‘í•©ë‹ˆë‹¤.")
        
        # ìƒˆ íƒ­ìœ¼ë¡œ ì „í™˜ (ì‚¬ìš©ìê°€ ìƒˆ íƒ­ì—ì„œ ìƒí’ˆ í˜ì´ì§€ë¡œ ì´ë™í–ˆì„ ê²½ìš°)
        self.update_log("ğŸ”„ í™œì„± íƒ­ì„ í™•ì¸í•˜ê³  ì „í™˜í•©ë‹ˆë‹¤...")
        try:
            # í˜„ì¬ íƒ­ ìˆ˜ í™•ì¸
            current_tabs = driver.window_handles
            self.update_log(f"ğŸ“‹ í˜„ì¬ ì—´ë¦° íƒ­ ìˆ˜: {len(current_tabs)}")
            
            if len(current_tabs) > 1:
                # ìƒˆ íƒ­ì´ ìˆë‹¤ë©´ ë§ˆì§€ë§‰ íƒ­(ìƒˆ íƒ­)ìœ¼ë¡œ ì „í™˜
                driver.switch_to.window(current_tabs[-1])
                self.update_log("âœ… ìƒˆ íƒ­ìœ¼ë¡œ ì „í™˜í–ˆìŠµë‹ˆë‹¤.")
                
                # í˜„ì¬ URL í™•ì¸
                current_url = driver.current_url
                self.update_log(f"ğŸ“ í˜„ì¬ í˜ì´ì§€ URL: {current_url}")
            else:
                self.update_log("â„¹ï¸ ìƒˆ íƒ­ì´ ì—†ìŠµë‹ˆë‹¤. í˜„ì¬ íƒ­ì„ ì‚¬ìš©í•©ë‹ˆë‹¤.")
                
        except Exception as e:
            self.update_log(f"âš ï¸ íƒ­ ì „í™˜ ì‹¤íŒ¨: {e}")
        
        time.sleep(sleep_time)

        # í˜ì´ì§€ ëê¹Œì§€ ìŠ¤í¬ë¡¤ (í˜„ì¬ í™œì„± íƒ­ì—ì„œ)
        self.update_log("ğŸ“œ í˜„ì¬ í™œì„± í˜ì´ì§€ í•˜ë‹¨ìœ¼ë¡œ ìŠ¤í¬ë¡¤í•©ë‹ˆë‹¤...")
        try:
            driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            self.update_log("âœ… í˜ì´ì§€ í•˜ë‹¨ìœ¼ë¡œ ìŠ¤í¬ë¡¤í–ˆìŠµë‹ˆë‹¤.")
        except Exception as e:
            self.update_log(f"âŒ ìŠ¤í¬ë¡¤ ì‹¤íŒ¨: {e}")
            self.close_alert_if_present(driver)

        time.sleep(sleep_time)

        # ë¦¬ë·° íƒ­ í´ë¦­ (JavaScript ì‚¬ìš©)
        self.update_log("ğŸ” ë¦¬ë·° íƒ­ì„ ì°¾ëŠ” ì¤‘...")
        try:
            product_review_button = driver.find_element(By.CSS_SELECTOR, product_review_button_selector)
            self.update_log("âœ… ë¦¬ë·° íƒ­ì„ ì°¾ì•˜ìŠµë‹ˆë‹¤.")
            # JavaScriptë¥¼ ì‚¬ìš©í•´ì„œ í´ë¦­ (ë‹¤ë¥¸ ìš”ì†Œì— ê°€ë ¤ì§„ ê²½ìš° í•´ê²°)
            driver.execute_script("arguments[0].click();", product_review_button)
            self.update_log("âœ… ë¦¬ë·° íƒ­ì„ í´ë¦­í–ˆìŠµë‹ˆë‹¤.")
        except Exception as e:
            self.update_log(f"âŒ ë¦¬ë·° íƒ­ í´ë¦­ ì‹¤íŒ¨: {e}")
            # ëŒ€ì•ˆ: JavaScriptë¡œ ì§ì ‘ í´ë¦­ ì‹œë„
            self.update_log("ğŸ”„ JavaScriptë¡œ ì§ì ‘ í´ë¦­ì„ ì‹œë„í•©ë‹ˆë‹¤...")
            try:
                driver.execute_script("document.querySelector('a[data-name=\"REVIEW\"]').click();")
                self.update_log("âœ… JavaScriptë¡œ ë¦¬ë·° íƒ­ì„ í´ë¦­í–ˆìŠµë‹ˆë‹¤.")
            except Exception as e2:
                self.update_log(f"âŒ JavaScript í´ë¦­ë„ ì‹¤íŒ¨: {e2}")
            self.close_alert_if_present(driver)

        time.sleep(3)

        # ë¦¬ë·° íƒ­ì´ ì œëŒ€ë¡œ í´ë¦­ë˜ì—ˆëŠ”ì§€ í™•ì¸
        self.update_log("ğŸ” ë¦¬ë·° íƒ­ í™œì„±í™” ìƒíƒœë¥¼ í™•ì¸í•©ë‹ˆë‹¤...")
        try:
            # ë¦¬ë·° íƒ­ì´ í™œì„±í™”ë˜ì—ˆëŠ”ì§€ í™•ì¸ (aria-current="true" ë˜ëŠ” í´ë˜ìŠ¤ ë³€ê²½ í™•ì¸)
            review_tab = driver.find_element(By.CSS_SELECTOR, 'a[data-name="REVIEW"]')
            if review_tab.get_attribute('aria-current') == 'true':
                self.update_log("âœ… ë¦¬ë·° íƒ­ì´ ì„±ê³µì ìœ¼ë¡œ í™œì„±í™”ë˜ì—ˆìŠµë‹ˆë‹¤.")
            else:
                self.update_log("âš ï¸ ë¦¬ë·° íƒ­ í™œì„±í™” ìƒíƒœë¥¼ í™•ì¸í•  ìˆ˜ ì—†ìŠµë‹ˆë‹¤.")
        except Exception as e:
            self.update_log(f"âŒ ë¦¬ë·° íƒ­ ìƒíƒœ í™•ì¸ ì‹¤íŒ¨: {e}")

        # ë¦¬ë·° íƒ­ í´ë¦­ í›„ ì •ë ¬ ì˜µì…˜ ì°¾ê¸° ë° í´ë¦­
        self.update_log("ğŸ”§ ì •ë ¬ ì˜µì…˜ì„ ì„¤ì •í•©ë‹ˆë‹¤...")
        try:
            # ì •ë ¬ ì˜µì…˜ë“¤ì„ ì°¾ê¸° ìœ„í•´ ì ì‹œ ëŒ€ê¸°
            self.update_log("â³ ì •ë ¬ ì˜µì…˜ ë¡œë”©ì„ ìœ„í•´ 3ì´ˆ ëŒ€ê¸°...")
            time.sleep(3)
            
            # ì •ë ¬ ì˜µì…˜ì— ë”°ë¥¸ ë²„íŠ¼ í´ë¦­ (ìƒˆë¡œìš´ êµ¬ì¡°)
            if self.sort_option == "latest":
                self.update_log("ğŸ“… ìµœì‹ ìˆœ ì •ë ¬ì„ ì‹œë„í•©ë‹ˆë‹¤...")
                # ìµœì‹ ìˆœ ì •ë ¬ ë²„íŠ¼ ì°¾ê¸°
                try:
                    new_review_button = driver.find_element(By.XPATH, "//a[contains(text(), 'ìµœì‹ ìˆœ')]")
                    self.update_log("âœ… ìµœì‹ ìˆœ ë²„íŠ¼ì„ ì°¾ì•˜ìŠµë‹ˆë‹¤.")
                    driver.execute_script("arguments[0].click();", new_review_button)
                    self.update_log("âœ… ìµœì‹ ìˆœ ì •ë ¬ ë²„íŠ¼ì„ í´ë¦­í–ˆìŠµë‹ˆë‹¤.")
                    time.sleep(2)  # ì •ë ¬ ë³€ê²½ í›„ ëŒ€ê¸°
                except:
                    self.update_log("ğŸ”„ ëŒ€ì•ˆ ì…€ë ‰í„°ë¡œ ìµœì‹ ìˆœ ë²„íŠ¼ì„ ì°¾ìŠµë‹ˆë‹¤...")
                    try:
                        new_review_button = driver.find_element(By.CSS_SELECTOR, 'a[data-shp-contents-id="ìµœì‹ ìˆœ"]')
                        driver.execute_script("arguments[0].click();", new_review_button)
                        self.update_log("âœ… ìµœì‹ ìˆœ ì •ë ¬ ë²„íŠ¼ì„ í´ë¦­í–ˆìŠµë‹ˆë‹¤.")
                        time.sleep(2)  # ì •ë ¬ ë³€ê²½ í›„ ëŒ€ê¸°
                    except Exception as e:
                        self.update_log(f"âŒ ìµœì‹ ìˆœ ì •ë ¬ ë²„íŠ¼ì„ ì°¾ì„ ìˆ˜ ì—†ìŠµë‹ˆë‹¤: {e}")
            elif self.sort_option == "lowest":
                self.update_log("â­ í‰ì  ë‚®ì€ìˆœ ì •ë ¬ì„ ì‹œë„í•©ë‹ˆë‹¤...")
                # í‰ì  ë‚®ì€ ìˆœ ì •ë ¬ ë²„íŠ¼ ì°¾ê¸°
                try:
                    bad_review_button = driver.find_element(By.XPATH, "//a[contains(text(), 'í‰ì  ë‚®ì€ìˆœ')]")
                    self.update_log("âœ… í‰ì  ë‚®ì€ìˆœ ë²„íŠ¼ì„ ì°¾ì•˜ìŠµë‹ˆë‹¤.")
                    driver.execute_script("arguments[0].click();", bad_review_button)
                    self.update_log("âœ… í‰ì  ë‚®ì€ ìˆœ ì •ë ¬ ë²„íŠ¼ì„ í´ë¦­í–ˆìŠµë‹ˆë‹¤.")
                    time.sleep(2)  # ì •ë ¬ ë³€ê²½ í›„ ëŒ€ê¸°
                except:
                    self.update_log("ğŸ”„ ëŒ€ì•ˆ ì…€ë ‰í„°ë¡œ í‰ì  ë‚®ì€ìˆœ ë²„íŠ¼ì„ ì°¾ìŠµë‹ˆë‹¤...")
                    try:
                        bad_review_button = driver.find_element(By.CSS_SELECTOR, 'a[data-shp-contents-id="í‰ì  ë‚®ì€ìˆœ"]')
                        driver.execute_script("arguments[0].click();", bad_review_button)
                        self.update_log("âœ… í‰ì  ë‚®ì€ ìˆœ ì •ë ¬ ë²„íŠ¼ì„ í´ë¦­í–ˆìŠµë‹ˆë‹¤.")
                        time.sleep(2)  # ì •ë ¬ ë³€ê²½ í›„ ëŒ€ê¸°
                    except Exception as e:
                        self.update_log(f"âŒ í‰ì  ë‚®ì€ ìˆœ ì •ë ¬ ë²„íŠ¼ì„ ì°¾ì„ ìˆ˜ ì—†ìŠµë‹ˆë‹¤: {e}")
            else:
                self.update_log(f"ğŸ“Š ê¸°ë³¸ ì •ë ¬ ì˜µì…˜ ì‚¬ìš©: {self.sort_option}")
        except Exception as e:
            self.update_log(f"âŒ ì •ë ¬ ì˜µì…˜ ì²˜ë¦¬ ì¤‘ ì˜¤ë¥˜: {e}")

        time.sleep(2)

        self.update_log("ğŸ“œ ì •ë ¬ í›„ í˜ì´ì§€ í•˜ë‹¨ìœ¼ë¡œ ìŠ¤í¬ë¡¤í•©ë‹ˆë‹¤...")
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        self.update_log("âœ… í˜ì´ì§€ í•˜ë‹¨ìœ¼ë¡œ ìŠ¤í¬ë¡¤í–ˆìŠµë‹ˆë‹¤.")

        # ë¦¬ë·° ì¶”ì¶œ (ë¦¬ë·° íƒ­ í´ë¦­ í›„ ìƒˆë¡œìš´ ì…€ë ‰í„° ì‚¬ìš©)
        self.update_log("ğŸ” ë¦¬ë·° ëª©ë¡ì˜ ì…€ë ‰í„°ë¥¼ ì°¾ëŠ” ì¤‘...")
        
        # ë¦¬ë·° ë¡œë”©ì„ ìœ„í•œ ì¶”ê°€ ëŒ€ê¸°
        self.update_log("â³ ë¦¬ë·° ëª©ë¡ ë¡œë”©ì„ ìœ„í•´ 5ì´ˆ ì¶”ê°€ ëŒ€ê¸°...")
        time.sleep(5)
        
        # í˜ì´ì§€ ì†ŒìŠ¤ ì¼ë¶€ë¥¼ í™•ì¸í•˜ì—¬ ë””ë²„ê¹… ì •ë³´ ì œê³µ
        try:
            page_source = driver.page_source
            if "ul.RR2FSL9wTc" in page_source:
                self.update_log("âœ… HTMLì—ì„œ ë¦¬ë·° ëª©ë¡ êµ¬ì¡°ë¥¼ í™•ì¸í–ˆìŠµë‹ˆë‹¤.")
            else:
                self.update_log("âš ï¸ HTMLì—ì„œ ë¦¬ë·° ëª©ë¡ êµ¬ì¡°ë¥¼ ì°¾ì„ ìˆ˜ ì—†ìŠµë‹ˆë‹¤.")
            
            if "li.PxsZltB5tV" in page_source:
                self.update_log("âœ… HTMLì—ì„œ ë¦¬ë·° ì•„ì´í…œ êµ¬ì¡°ë¥¼ í™•ì¸í–ˆìŠµë‹ˆë‹¤.")
            else:
                self.update_log("âš ï¸ HTMLì—ì„œ ë¦¬ë·° ì•„ì´í…œ êµ¬ì¡°ë¥¼ ì°¾ì„ ìˆ˜ ì—†ìŠµë‹ˆë‹¤.")
                
            # ìƒˆë¡œìš´ êµ¬ì¡° í™•ì¸
            if "ul.HTT4L8U0CU" in page_source:
                self.update_log("âœ… HTMLì—ì„œ ìƒˆë¡œìš´ ë¦¬ë·° ëª©ë¡ êµ¬ì¡°ë¥¼ í™•ì¸í–ˆìŠµë‹ˆë‹¤.")
            else:
                self.update_log("âš ï¸ HTMLì—ì„œ ìƒˆë¡œìš´ ë¦¬ë·° ëª©ë¡ êµ¬ì¡°ë¥¼ ì°¾ì„ ìˆ˜ ì—†ìŠµë‹ˆë‹¤.")
                
        except Exception as e:
            self.update_log(f"âš ï¸ í˜ì´ì§€ ì†ŒìŠ¤ í™•ì¸ ì‹¤íŒ¨: {e}")
        
        # ë¦¬ë·° ëª©ë¡ì˜ ì…€ë ‰í„°ë¥¼ ë™ì ìœ¼ë¡œ ì°¾ê¸°
        review_selector = self.find_review_selector(driver)
        if review_selector:
            self.update_log("âœ… ë¦¬ë·° ì…€ë ‰í„°ë¥¼ ì°¾ì•˜ìŠµë‹ˆë‹¤. ë¦¬ë·° í¬ë¡¤ë§ì„ ì‹œì‘í•©ë‹ˆë‹¤.")
            self.crawl_reviews_until_page(driver, last_page_to_crawl, self.folder_path, review_selector)
        else:
            self.update_log("âŒ ë¦¬ë·° ì…€ë ‰í„°ë¥¼ ì°¾ì„ ìˆ˜ ì—†ìŠµë‹ˆë‹¤.")
            # ì¶”ê°€ ë””ë²„ê¹… ì •ë³´ ì œê³µ
            self.update_log("ğŸ” ë””ë²„ê¹…ì„ ìœ„í•´ í˜ì´ì§€ êµ¬ì¡°ë¥¼ í™•ì¸í•©ë‹ˆë‹¤...")
            try:
                # REVIEW ì„¹ì…˜ì˜ ëª¨ë“  ìš”ì†Œ í™•ì¸
                review_section = driver.find_element(By.CSS_SELECTOR, "#REVIEW")
                self.update_log(f"âœ… REVIEW ì„¹ì…˜ ë°œê²¬: {review_section.tag_name}")
                
                # ëª¨ë“  ul ìš”ì†Œ í™•ì¸
                ul_elements = driver.find_elements(By.CSS_SELECTOR, "#REVIEW ul")
                self.update_log(f"ğŸ“‹ REVIEW ì„¹ì…˜ ë‚´ ul ìš”ì†Œ ìˆ˜: {len(ul_elements)}")
                
                for i, ul in enumerate(ul_elements):
                    ul_class = ul.get_attribute('class')
                    li_count = len(ul.find_elements(By.CSS_SELECTOR, 'li'))
                    self.update_log(f"  ul {i+1}: class='{ul_class}', li ê°œìˆ˜={li_count}")
                    
                    # ë¦¬ë·° ëª©ë¡ì¸ì§€ í™•ì¸ (ë§ì€ li ìš”ì†Œë¥¼ ê°€ì§„ ul)
                    if li_count > 10:  # ë¦¬ë·° ëª©ë¡ì€ ë³´í†µ 10ê°œ ì´ìƒì˜ lië¥¼ ê°€ì§
                        self.update_log(f"    ğŸ¯ ì ì¬ì  ë¦¬ë·° ëª©ë¡ ë°œê²¬: {ul_class}")
                        # ì´ ulì˜ lië“¤ì„ í™•ì¸
                        li_elements_in_ul = ul.find_elements(By.CSS_SELECTOR, 'li')
                        for j, li in enumerate(li_elements_in_ul[:3]):  # ì²˜ìŒ 3ê°œë§Œ í™•ì¸
                            li_class = li.get_attribute('class')
                            li_html = li.get_attribute('outerHTML')[:100]
                            self.update_log(f"      li {j+1}: class='{li_class}' (HTML: {li_html}...)")
                
                # ëª¨ë“  li ìš”ì†Œ í™•ì¸
                li_elements = driver.find_elements(By.CSS_SELECTOR, "#REVIEW li")
                self.update_log(f"ğŸ“‹ REVIEW ì„¹ì…˜ ë‚´ li ìš”ì†Œ ìˆ˜: {len(li_elements)}")
                
                for i, li in enumerate(li_elements[:5]):  # ì²˜ìŒ 5ê°œë§Œ í™•ì¸
                    li_class = li.get_attribute('class')
                    self.update_log(f"  li {i+1}: class='{li_class}'")
                    
            except Exception as e:
                self.update_log(f"âŒ ë””ë²„ê¹… ì •ë³´ ìˆ˜ì§‘ ì‹¤íŒ¨: {e}")

        # ë“œë¼ì´ë²„ ì¢…ë£Œ
        self.update_log("ğŸ”š Chrome ë“œë¼ì´ë²„ë¥¼ ì¢…ë£Œí•©ë‹ˆë‹¤...")
        driver.quit()
        self.update_log("âœ… Chrome ë“œë¼ì´ë²„ ì¢…ë£Œ ì™„ë£Œ")

        self.update_log("=" * 50)
        self.update_log("ğŸ‰ í¬ë¡¤ë§ì´ ì™„ë£Œë˜ì—ˆìŠµë‹ˆë‹¤!")
        self.update_log("=" * 50)

    def close_alert_if_present(self, driver):
        try:
            alert = driver.switch_to.alert
            alert.dismiss()
            self.update_log("ì˜ˆê¸°ì¹˜ ì•Šì€ ì•Œë¦¼ì°½ì„ ë‹«ì•˜ìŠµë‹ˆë‹¤.")
        except:
            pass

    def find_review_selector(self, driver):
        """ë¦¬ë·° ëª©ë¡ì˜ ì…€ë ‰í„°ë¥¼ ë™ì ìœ¼ë¡œ ì°¾ê¸° (ìƒˆë¡œìš´ êµ¬ì¡°)"""
        # ë¦¬ë·° ì„¹ì…˜ì´ ë¡œë“œë  ë•Œê¹Œì§€ ëŒ€ê¸°
        try:
            WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "#REVIEW"))
            )
            self.update_log("ë¦¬ë·° ì„¹ì…˜ì„ ì°¾ì•˜ìŠµë‹ˆë‹¤.")
        except:
            self.update_log("ë¦¬ë·° ì„¹ì…˜ì„ ì°¾ì„ ìˆ˜ ì—†ìŠµë‹ˆë‹¤.")
        
        # HTML êµ¬ì¡°ë¥¼ ë” ì •í™•íˆ ë¶„ì„í•˜ì—¬ ì…€ë ‰í„° ìš°ì„ ìˆœìœ„ ì¡°ì •
        possible_selectors = [
            # ê°€ì¥ ì •í™•í•œ ì…€ë ‰í„°ë“¤ (HTML êµ¬ì¡° ê¸°ë°˜)
            "ul.RR2FSL9wTc li.PxsZltB5tV",  # ë©”ì¸ ë¦¬ë·° ëª©ë¡
            "li.PxsZltB5tV",  # ë¦¬ë·° ì•„ì´í…œë“¤
            "#REVIEW ul.RR2FSL9wTc li",  # ë¦¬ë·° ì„¹ì…˜ ë‚´ì˜ ì •í™•í•œ ë¦¬ë·° ëª©ë¡
            # ìƒˆë¡œìš´ êµ¬ì¡° ì‹œë„ (ë¡œê·¸ì—ì„œ í™•ì¸ëœ êµ¬ì¡°)
            "ul.HTT4L8U0CU li",  # ìƒˆë¡œìš´ ë¦¬ë·° ëª©ë¡ êµ¬ì¡°
            "#REVIEW ul.HTT4L8U0CU li",  # ë¦¬ë·° ì„¹ì…˜ ë‚´ì˜ ìƒˆë¡œìš´ ë¦¬ë·° ëª©ë¡
            # ë²”ìš©ì ì¸ ì…€ë ‰í„°ë“¤
            "#REVIEW ul li",  # ë¦¬ë·° ì„¹ì…˜ ë‚´ì˜ li ìš”ì†Œë“¤
            "#REVIEW li",  # ë¦¬ë·° ì„¹ì…˜ ë‚´ì˜ ëª¨ë“  li
            # ì¶”ê°€ì ì¸ ì…€ë ‰í„°ë“¤
            "li[class*='PxsZltB5tV']",
            "ul[class*='RR2FSL9wTc'] li",
            "li[class*='PxsZltB5tV'][class*='_nlog_click']",
            ".review_list li",
            "[data-testid='review-item']",
            ".review-item",
            "ul li[class*='review']",
            "li[class*='review']",
            "div[class*='review'] li"
        ]
        
        for selector in possible_selectors:
            try:
                elements = driver.find_elements(By.CSS_SELECTOR, selector)
                self.update_log(f"ì…€ë ‰í„° '{selector}' í…ŒìŠ¤íŠ¸: {len(elements)}ê°œ ìš”ì†Œ ë°œê²¬")
                if elements and len(elements) > 0:
                    # ì‹¤ì œ ë¦¬ë·°ì¸ì§€ í™•ì¸ (ë¦¬ë·° ì ìˆ˜ê°€ ìˆëŠ”ì§€ ì²´í¬)
                    valid_reviews = 0
                    for i, elem in enumerate(elements[:3]):  # ì²˜ìŒ 3ê°œë§Œ í™•ì¸
                        try:
                            # ë¦¬ë·° ì ìˆ˜ ìš”ì†Œê°€ ìˆëŠ”ì§€ í™•ì¸ (HTML êµ¬ì¡°ì— ë§ëŠ” ì…€ë ‰í„° ìš°ì„ )
                            score_elem = None
                            score_selectors = [
                                'em.n6zq2yy0KA',  # HTMLì—ì„œ í™•ì¸ëœ ì •í™•í•œ í´ë˜ìŠ¤
                                'div.AlfkEF45qI em.n6zq2yy0KA',  # ë” êµ¬ì²´ì ì¸ ê²½ë¡œ
                                'div.AlfkEF45qI em',
                                'em[class*="score"]',
                                'em[class*="star"]',
                                'em[class*="rating"]',
                                'em',
                                'span[class*="score"]',
                                'span[class*="star"]'
                            ]
                            
                            for score_selector in score_selectors:
                                try:
                                    score_elem = elem.find_element(By.CSS_SELECTOR, score_selector)
                                    if score_elem and score_elem.text and score_elem.text.strip().isdigit():
                                        break
                                except:
                                    continue
                            
                            if score_elem and score_elem.text and score_elem.text.strip().isdigit():
                                valid_reviews += 1
                                self.update_log(f"  ìš”ì†Œ {i+1}: ì ìˆ˜ '{score_elem.text}' ë°œê²¬")
                            else:
                                # ì ìˆ˜ ìš”ì†Œë¥¼ ì°¾ì§€ ëª»í•œ ê²½ìš°, ìš”ì†Œì˜ HTML êµ¬ì¡°ë¥¼ í™•ì¸
                                try:
                                    elem_html = elem.get_attribute('outerHTML')[:200]  # ì²˜ìŒ 200ìë§Œ
                                    self.update_log(f"  ìš”ì†Œ {i+1}: ì ìˆ˜ ìš”ì†Œ ì—†ìŒ (HTML: {elem_html}...)")
                                except:
                                    self.update_log(f"  ìš”ì†Œ {i+1}: ì ìˆ˜ ìš”ì†Œ ì—†ìŒ")
                        except Exception as e:
                            self.update_log(f"  ìš”ì†Œ {i+1}: ì ìˆ˜ í™•ì¸ ì‹¤íŒ¨ - {e}")
                            continue
                    
                    if valid_reviews > 0:
                        self.update_log(f"âœ… ë¦¬ë·° ì…€ë ‰í„° ë°œê²¬: {selector} ({len(elements)}ê°œ ë¦¬ë·°, {valid_reviews}ê°œ ìœ íš¨)")
                        return selector
                    else:
                        self.update_log(f"âŒ ì…€ë ‰í„° '{selector}': ìœ íš¨í•œ ë¦¬ë·° ì—†ìŒ")
            except Exception as e:
                self.update_log(f"âŒ ì…€ë ‰í„° '{selector}' í…ŒìŠ¤íŠ¸ ì‹¤íŒ¨: {e}")
                continue
        
        # XPathë¡œë„ ì‹œë„
        xpath_selectors = [
            "//ul[contains(@class, 'RR2FSL9wTc')]/li[contains(@class, 'PxsZltB5tV')]",
            "//div[@id='REVIEW']//ul[contains(@class, 'RR2FSL9wTc')]//li",
            # ìƒˆë¡œìš´ êµ¬ì¡° ì‹œë„
            "//ul[contains(@class, 'HTT4L8U0CU')]//li",
            "//div[@id='REVIEW']//ul[contains(@class, 'HTT4L8U0CU')]//li",
            # ë²”ìš©ì ì¸ XPath
            "//div[@id='REVIEW']//ul//li",
            "//li[contains(@class, 'PxsZltB5tV')]"
        ]
        
        for xpath in xpath_selectors:
            try:
                elements = driver.find_elements(By.XPATH, xpath)
                self.update_log(f"XPath '{xpath}' í…ŒìŠ¤íŠ¸: {len(elements)}ê°œ ìš”ì†Œ ë°œê²¬")
                
                if elements and len(elements) > 0:
                    # ì‹¤ì œ ë¦¬ë·°ì¸ì§€ í™•ì¸
                    valid_reviews = 0
                    for i, elem in enumerate(elements[:3]):
                        try:
                            # ë¦¬ë·° ì ìˆ˜ ìš”ì†Œê°€ ìˆëŠ”ì§€ í™•ì¸ (HTML êµ¬ì¡°ì— ë§ëŠ” ì…€ë ‰í„° ìš°ì„ )
                            score_elem = None
                            score_selectors = [
                                'em.n6zq2yy0KA',  # HTMLì—ì„œ í™•ì¸ëœ ì •í™•í•œ í´ë˜ìŠ¤
                                'div.AlfkEF45qI em.n6zq2yy0KA',  # ë” êµ¬ì²´ì ì¸ ê²½ë¡œ
                                'div.AlfkEF45qI em',
                                'em[class*="score"]',
                                'em[class*="star"]',
                                'em[class*="rating"]',
                                'em',
                                'span[class*="score"]',
                                'span[class*="star"]'
                            ]
                            
                            for score_selector in score_selectors:
                                try:
                                    score_elem = elem.find_element(By.CSS_SELECTOR, score_selector)
                                    if score_elem and score_elem.text and score_elem.text.strip().isdigit():
                                        break
                                except:
                                    continue
                            
                            if score_elem and score_elem.text and score_elem.text.strip().isdigit():
                                valid_reviews += 1
                                self.update_log(f"  ìš”ì†Œ {i+1}: ì ìˆ˜ '{score_elem.text}' ë°œê²¬")
                            else:
                                # ì ìˆ˜ ìš”ì†Œë¥¼ ì°¾ì§€ ëª»í•œ ê²½ìš°, ìš”ì†Œì˜ HTML êµ¬ì¡°ë¥¼ í™•ì¸
                                try:
                                    elem_html = elem.get_attribute('outerHTML')[:200]  # ì²˜ìŒ 200ìë§Œ
                                    self.update_log(f"  ìš”ì†Œ {i+1}: ì ìˆ˜ ìš”ì†Œ ì—†ìŒ (HTML: {elem_html}...)")
                                except:
                                    self.update_log(f"  ìš”ì†Œ {i+1}: ì ìˆ˜ ìš”ì†Œ ì—†ìŒ")
                        except Exception as e:
                            self.update_log(f"  ìš”ì†Œ {i+1}: ì ìˆ˜ í™•ì¸ ì‹¤íŒ¨ - {e}")
                            continue
                    
                    if valid_reviews > 0:
                        self.update_log(f"âœ… ë¦¬ë·° ì…€ë ‰í„° ë°œê²¬ (XPath): {xpath} ({len(elements)}ê°œ ë¦¬ë·°, {valid_reviews}ê°œ ìœ íš¨)")
                        return xpath
                    else:
                        self.update_log(f"âŒ XPath '{xpath}': ìœ íš¨í•œ ë¦¬ë·° ì—†ìŒ")
            except Exception as e:
                self.update_log(f"âŒ XPath '{xpath}' í…ŒìŠ¤íŠ¸ ì‹¤íŒ¨: {e}")
                continue
            
        # ë§ˆì§€ë§‰ ì‹œë„: ë¦¬ë·°ê°€ ë¡œë“œë  ë•Œê¹Œì§€ ëŒ€ê¸°í•˜ê³  ë‹¤ì‹œ ì‹œë„
        self.update_log("ğŸ”„ ë¦¬ë·° ë¡œë”©ì„ ìœ„í•´ ì¶”ê°€ ëŒ€ê¸° í›„ ì¬ì‹œë„...")
        time.sleep(10)  # 10ì´ˆ ì¶”ê°€ ëŒ€ê¸°
        
        # ë‹¤ì‹œ ì‹œë„
        for selector in possible_selectors[:5]:  # ìƒìœ„ 5ê°œ ì…€ë ‰í„°ë§Œ ì¬ì‹œë„
            try:
                elements = driver.find_elements(By.CSS_SELECTOR, selector)
                self.update_log(f"ì¬ì‹œë„ - ì…€ë ‰í„° '{selector}' í…ŒìŠ¤íŠ¸: {len(elements)}ê°œ ìš”ì†Œ ë°œê²¬")
                if elements and len(elements) > 0:
                    # ì‹¤ì œ ë¦¬ë·°ì¸ì§€ í™•ì¸
                    valid_reviews = 0
                    for i, elem in enumerate(elements[:3]):
                        try:
                            score_elem = None
                            score_selectors = [
                                'em.n6zq2yy0KA',
                                'div.AlfkEF45qI em.n6zq2yy0KA',
                                'div.AlfkEF45qI em',
                                'em[class*="score"]',
                                'em[class*="star"]',
                                'em[class*="rating"]',
                                'em',
                                'span[class*="score"]',
                                'span[class*="star"]'
                            ]
                            
                            for score_selector in score_selectors:
                                try:
                                    score_elem = elem.find_element(By.CSS_SELECTOR, score_selector)
                                    if score_elem and score_elem.text and score_elem.text.strip().isdigit():
                                        break
                                except:
                                    continue
                            
                            if score_elem and score_elem.text and score_elem.text.strip().isdigit():
                                valid_reviews += 1
                                self.update_log(f"  ì¬ì‹œë„ - ìš”ì†Œ {i+1}: ì ìˆ˜ '{score_elem.text}' ë°œê²¬")
                            else:
                                self.update_log(f"  ì¬ì‹œë„ - ìš”ì†Œ {i+1}: ì ìˆ˜ ìš”ì†Œ ì—†ìŒ")
                        except Exception as e:
                            self.update_log(f"  ì¬ì‹œë„ - ìš”ì†Œ {i+1}: ì ìˆ˜ í™•ì¸ ì‹¤íŒ¨ - {e}")
                            continue
                    
                    if valid_reviews > 0:
                        self.update_log(f"âœ… ì¬ì‹œë„ ì„±ê³µ - ë¦¬ë·° ì…€ë ‰í„° ë°œê²¬: {selector} ({len(elements)}ê°œ ë¦¬ë·°, {valid_reviews}ê°œ ìœ íš¨)")
                        return selector
            except Exception as e:
                self.update_log(f"âŒ ì¬ì‹œë„ - ì…€ë ‰í„° '{selector}' í…ŒìŠ¤íŠ¸ ì‹¤íŒ¨: {e}")
                continue
        
        self.update_log("ë¦¬ë·° ì…€ë ‰í„°ë¥¼ ì°¾ì„ ìˆ˜ ì—†ìŠµë‹ˆë‹¤.")
        return None

    def crawl_reviews_until_page(self, driver, last_page, photo_folder_path, review_selector):
        self.update_log("=" * 50)
        self.update_log("ğŸ“š ë¦¬ë·° í˜ì´ì§€ í¬ë¡¤ë§ì„ ì‹œì‘í•©ë‹ˆë‹¤...")
        self.update_log("=" * 50)
        current_page = 1
        chunk_count = 1
        while True:
            self.update_log(f"ğŸ“„ í˜ì´ì§€ {current_page} í¬ë¡¤ë§ ì‹œì‘...")
            chunk_reviews = []
            chunk_start_page = current_page
            
            for _ in range(self.chunk_size):
                if (last_page != "max" and current_page > last_page):
                    self.update_log(f"ğŸ“Š ì„¤ì •ëœ í˜ì´ì§€ ìˆ˜({last_page})ì— ë„ë‹¬í–ˆìŠµë‹ˆë‹¤.")
                    break
                
                self.update_log(f"ğŸ” í˜ì´ì§€ {current_page}ì—ì„œ ë¦¬ë·°ë¥¼ ì¶”ì¶œí•©ë‹ˆë‹¤...")
                reviews_data = self.extract_reviews(driver, photo_folder_path, current_page, review_selector)
                chunk_reviews.extend(reviews_data)
                self.update_log(f"ğŸ“Š í˜ì´ì§€ {current_page}ì—ì„œ {len(reviews_data)}ê°œ ë¦¬ë·° ì¶”ì¶œ ì™„ë£Œ")
                
                if not reviews_data:  # ë” ì´ìƒ ë¦¬ë·°ê°€ ì—†ìœ¼ë©´ ì¢…ë£Œ
                    self.update_log("âš ï¸ ë” ì´ìƒ ë¦¬ë·°ê°€ ì—†ìŠµë‹ˆë‹¤. í¬ë¡¤ë§ì„ ì¢…ë£Œí•©ë‹ˆë‹¤.")
                    break
                
                current_page += 1
                
                try:
                    self.update_log(f"ğŸ” í˜ì´ì§€ {current_page}ë¡œ ì´ë™ì„ ì‹œë„í•©ë‹ˆë‹¤...")
                    # í˜ì´ì§€ë„¤ì´ì…˜ ë²„íŠ¼ë“¤ ì°¾ê¸°
                    pagination_selectors = [
                        'a.hyY6CXtbcn[aria-current="true"]',  # í˜„ì¬ í˜ì´ì§€
                        'a.hyY6CXtbcn',  # í˜ì´ì§€ ë²ˆí˜¸ë“¤
                        'a.JY2WGJ4hXh.I3i1NSoFdB',  # ë‹¤ìŒ ë²„íŠ¼
                        'a[aria-label="ë‹¤ìŒ í˜ì´ì§€"]',  # ë‹¤ìŒ ë²„íŠ¼ ëŒ€ì•ˆ
                        'a[title="ë‹¤ìŒ"]'  # ë‹¤ìŒ ë²„íŠ¼ ëŒ€ì•ˆ
                    ]
                    
                    # í˜„ì¬ í˜ì´ì§€ í™•ì¸
                    self.update_log(f"  ğŸ“ í˜„ì¬ í˜ì´ì§€ë¥¼ í™•ì¸í•©ë‹ˆë‹¤...")
                    current_page_element = None
                    for selector in pagination_selectors[:1]:  # í˜„ì¬ í˜ì´ì§€ë§Œ í™•ì¸
                        try:
                            current_page_element = driver.find_element(By.CSS_SELECTOR, selector)
                            break
                        except:
                            continue
                    
                    if current_page_element:
                        current_page_number = int(current_page_element.text)
                        self.update_log(f"  âœ… í˜„ì¬ í˜ì´ì§€: {current_page_number}")
                    else:
                        self.update_log(f"  âš ï¸ í˜„ì¬ í˜ì´ì§€ë¥¼ í™•ì¸í•  ìˆ˜ ì—†ìŠµë‹ˆë‹¤.")
                    
                    # ë‹¤ìŒ í˜ì´ì§€ë¡œ ì´ë™ ì‹œë„
                    self.update_log(f"  ğŸ”„ ë‹¤ìŒ í˜ì´ì§€ë¡œ ì´ë™ì„ ì‹œë„í•©ë‹ˆë‹¤...")
                    next_page_clicked = False
                    
                    # 1. ë‹¤ìŒ í˜ì´ì§€ ë²ˆí˜¸ ë²„íŠ¼ í´ë¦­ ì‹œë„
                    self.update_log(f"    ğŸ“„ í˜ì´ì§€ ë²ˆí˜¸ ë²„íŠ¼ì„ ì°¾ëŠ” ì¤‘...")
                    try:
                        next_page_elements = driver.find_elements(By.CSS_SELECTOR, 'a.hyY6CXtbcn[aria-current="false"]')
                        for elem in next_page_elements:
                            if elem.text.isdigit() and int(elem.text) == current_page:
                                driver.execute_script("arguments[0].click();", elem)
                                next_page_clicked = True
                                self.update_log(f"    âœ… í˜ì´ì§€ {current_page}ë¡œ ì´ë™í–ˆìŠµë‹ˆë‹¤.")
                                break
                    except Exception as e:
                        self.update_log(f"    âŒ í˜ì´ì§€ ë²ˆí˜¸ í´ë¦­ ì‹¤íŒ¨: {e}")
                    
                    # 2. ë‹¤ìŒ ë²„íŠ¼ í´ë¦­ ì‹œë„
                    if not next_page_clicked:
                        self.update_log(f"    ğŸ”„ ë‹¤ìŒ ë²„íŠ¼ì„ ì°¾ëŠ” ì¤‘...")
                        for selector in pagination_selectors[2:]:  # ë‹¤ìŒ ë²„íŠ¼ë“¤ë§Œ í™•ì¸
                            try:
                                next_button = driver.find_element(By.CSS_SELECTOR, selector)
                                driver.execute_script("arguments[0].click();", next_button)
                                next_page_clicked = True
                                self.update_log("    âœ… ë‹¤ìŒ ë²„íŠ¼ì„ í´ë¦­í–ˆìŠµë‹ˆë‹¤.")
                                break
                            except:
                                continue
                    
                    if not next_page_clicked:
                        self.update_log("    âŒ ë‹¤ìŒ í˜ì´ì§€ë¡œ ì´ë™í•  ìˆ˜ ì—†ìŠµë‹ˆë‹¤.")
                        break
                    
                    self.update_log(f"    â³ í˜ì´ì§€ ë¡œë”©ì„ ìœ„í•´ {random.uniform(2, 4):.1f}ì´ˆ ëŒ€ê¸°...")
                    time.sleep(random.uniform(2, 4))  # í˜ì´ì§€ ë¡œë”© ëŒ€ê¸°
                    
                except Exception as e:
                    self.update_log(f"âŒ í˜ì´ì§€ ì´ë™ ì‹¤íŒ¨: {e}")
                    break
            
            # ì²­í¬ ë°ì´í„° ì €ì¥
            if chunk_reviews:
                self.update_log(f"ğŸ’¾ ì²­í¬ {chunk_count} ë°ì´í„°ë¥¼ ì €ì¥í•©ë‹ˆë‹¤...")
                excel_path = os.path.join(self.folder_path, f'reviews_chunk_{chunk_count}.xlsx')
                self.save_to_excel(chunk_reviews, excel_path, chunk_start_page)
                self.update_log(f"âœ… ì²­í¬ {chunk_count} ì €ì¥ ì™„ë£Œ: {len(chunk_reviews)}ê°œ ë¦¬ë·°")
                chunk_count += 1
            else:
                self.update_log("âš ï¸ ì €ì¥í•  ë¦¬ë·° ë°ì´í„°ê°€ ì—†ìŠµë‹ˆë‹¤.")
            
            if (last_page != "max" and current_page > last_page) or not reviews_data:
                self.update_log("ğŸ í¬ë¡¤ë§ ì¡°ê±´ì„ ë§Œì¡±í•˜ì—¬ ì¢…ë£Œí•©ë‹ˆë‹¤.")
                break

    def extract_reviews(self, driver, photo_folder_path, current_page, review_selector):
        self.update_log(f"ğŸ“„ í˜ì´ì§€ {current_page}ì—ì„œ ë¦¬ë·°ë¥¼ ì¶”ì¶œí•©ë‹ˆë‹¤...")
        reviews_data = []
        try:
            # CSS ì…€ë ‰í„°ì¸ì§€ XPathì¸ì§€ í™•ì¸
            if review_selector.startswith("//"):
                self.update_log(f"ğŸ” XPathë¡œ ë¦¬ë·°ë¥¼ ì°¾ëŠ” ì¤‘: {review_selector}")
                reviews = driver.find_elements(By.XPATH, review_selector)
            else:
                self.update_log(f"ğŸ” CSS ì…€ë ‰í„°ë¡œ ë¦¬ë·°ë¥¼ ì°¾ëŠ” ì¤‘: {review_selector}")
                reviews = driver.find_elements(By.CSS_SELECTOR, review_selector)
            
            self.update_log(f"âœ… ë¦¬ë·° {len(reviews)}ê°œë¥¼ ì°¾ì•˜ìŠµë‹ˆë‹¤.")
            
            if not reviews:
                self.update_log("ë¦¬ë·°ë¥¼ ì°¾ì„ ìˆ˜ ì—†ìŠµë‹ˆë‹¤. í˜ì´ì§€ êµ¬ì¡°ê°€ ë³€ê²½ë˜ì—ˆì„ ìˆ˜ ìˆìŠµë‹ˆë‹¤.")
                return reviews_data
            
            for index, review in enumerate(reviews):
                self.update_log(f"ğŸ“ ë¦¬ë·° {index+1}/{len(reviews)} ì¶”ì¶œ ì¤‘...")
                try:
                    # ë¦¬ë·° ì ìˆ˜ ì¶”ì¶œ (HTML êµ¬ì¡°ì— ë§ëŠ” ì…€ë ‰í„° ìš°ì„ )
                    self.update_log(f"  â­ ë¦¬ë·° ì ìˆ˜ ì¶”ì¶œ ì¤‘...")
                    review_score = self.safe_extract_text(review, [
                        'em.n6zq2yy0KA',  # HTMLì—ì„œ í™•ì¸ëœ ì •í™•í•œ í´ë˜ìŠ¤
                        'div.AlfkEF45qI em.n6zq2yy0KA',  # ë” êµ¬ì²´ì ì¸ ê²½ë¡œ
                        'div.AlfkEF45qI em',
                        'em[class*="score"]',
                        'em[class*="star"]',
                        'em[class*="rating"]',
                        'em',
                        '.score',
                        'span[class*="score"]',
                        'span[class*="star"]'
                    ], "ì ìˆ˜ ì—†ìŒ")
                    self.update_log(f"  âœ… ë¦¬ë·° ì ìˆ˜: {review_score}")
                    
                    # ë¦¬ë·°ì–´ ì´ë¦„ ì¶”ì¶œ (HTML êµ¬ì¡°ì— ë§ëŠ” ì…€ë ‰í„° ìš°ì„ )
                    self.update_log(f"  ğŸ‘¤ ë¦¬ë·°ì–´ ì´ë¦„ ì¶”ì¶œ ì¤‘...")
                    reviewer_name = self.safe_extract_text(review, [
                        'strong.MX91DFZo2F',  # HTMLì—ì„œ í™•ì¸ëœ ì •í™•í•œ í´ë˜ìŠ¤
                        'div.AlfkEF45qI strong.MX91DFZo2F',  # ë” êµ¬ì²´ì ì¸ ê²½ë¡œ
                        'div.Db9Dtnf7gY strong.MX91DFZo2F',
                        'strong[class*="name"]',
                        '.reviewer-name',
                        'span[class*="name"]'
                    ], "ì´ë¦„ ì—†ìŒ")
                    self.update_log(f"  âœ… ë¦¬ë·°ì–´: {reviewer_name}")
                    
                    # ë¦¬ë·° ë‚ ì§œ ì¶”ì¶œ (HTML êµ¬ì¡°ì— ë§ëŠ” ì…€ë ‰í„° ìš°ì„ )
                    self.update_log(f"  ğŸ“… ë¦¬ë·° ë‚ ì§œ ì¶”ì¶œ ì¤‘...")
                    review_date = self.safe_extract_text(review, [
                        'span.MX91DFZo2F',  # HTMLì—ì„œ í™•ì¸ëœ ì •í™•í•œ í´ë˜ìŠ¤
                        'div.Db9Dtnf7gY span.MX91DFZo2F',  # ë” êµ¬ì²´ì ì¸ ê²½ë¡œ
                        'div.AlfkEF45qI div.Db9Dtnf7gY span.MX91DFZo2F',
                        'span[class*="date"]',
                        '.review-date',
                        'div[class*="date"]'
                    ], "ë‚ ì§œ ì—†ìŒ")
                    self.update_log(f"  âœ… ë¦¬ë·° ë‚ ì§œ: {review_date}")
                    
                    # ìƒí’ˆ ì˜µì…˜ ì •ë³´ ì¶”ì¶œ (HTML êµ¬ì¡°ì— ë§ëŠ” ì…€ë ‰í„° ìš°ì„ )
                    self.update_log(f"  ğŸ›ï¸ ìƒí’ˆ ì˜µì…˜ ì •ë³´ ì¶”ì¶œ ì¤‘...")
                    product_name = self.safe_extract_text(review, [
                        'div.b_caIle8kC',  # HTMLì—ì„œ í™•ì¸ëœ ì •í™•í•œ í´ë˜ìŠ¤
                        'div.AlfkEF45qI div.b_caIle8kC',  # ë” êµ¬ì²´ì ì¸ ê²½ë¡œ
                        'div[class*="product"]',
                        '.product-name',
                        'span[class*="product"]'
                    ], "ì •ë³´ ì—†ìŒ")
                    self.update_log(f"  âœ… ìƒí’ˆ ì˜µì…˜: {product_name}")

                    # ë¦¬ë·° ë‚´ìš© ì¶”ì¶œ (ì—¬ëŸ¬ ë°©ë²• ì‹œë„)
                    self.update_log(f"  ğŸ“ ë¦¬ë·° ë‚´ìš© ì¶”ì¶œ ì¤‘...")
                    content = self.safe_extract_review_content(review)
                    self.update_log(f"  âœ… ë¦¬ë·° ë‚´ìš© ê¸¸ì´: {len(content)}ì")

                    # ë¦¬ë·° ìœ í˜• í™•ì¸
                    self.update_log(f"  ğŸ·ï¸ ë¦¬ë·° ìœ í˜• í™•ì¸ ì¤‘...")
                    review_type = self.determine_review_type(review)
                    self.update_log(f"  âœ… ë¦¬ë·° ìœ í˜•: {review_type}")

                    # ë¦¬ë·° ì´ë¯¸ì§€ ì¶”ì¶œ
                    self.update_log(f"  ğŸ“¸ ë¦¬ë·° ì´ë¯¸ì§€ ì¶”ì¶œ ì¤‘...")
                    photos = self.extract_review_photos(review, photo_folder_path, current_page, index)
                    self.update_log(f"  âœ… ì¶”ì¶œëœ ì´ë¯¸ì§€ ìˆ˜: {len(photos)}ê°œ")

                    reviews_data.append({
                        "Review Score": review_score,
                        "Reviewer Name": reviewer_name,
                        "Review Date": review_date,
                        "Product(Option) Name": product_name,
                        "Review Type": review_type,
                        "Content": content,
                        "Photos": photos
                    })
                    self.update_log(f"  âœ… ë¦¬ë·° ì¶”ì¶œ ì™„ë£Œ: í˜ì´ì§€ {current_page}, ë¦¬ë·° {index+1}")
                    
                except Exception as e:
                    self.update_log(f"  âŒ ë¦¬ë·° {index+1} ì¶”ì¶œ ì‹¤íŒ¨: {e}")
                    continue
                    
        except Exception as e:
            self.update_log(f"ë¦¬ë·° ì •ë³´ ì¶”ì¶œ ì‹¤íŒ¨: {e}")
        return reviews_data

    def safe_extract_text(self, element, selectors, default_value=""):
        """ì—¬ëŸ¬ ì…€ë ‰í„°ë¥¼ ì‹œë„í•˜ì—¬ í…ìŠ¤íŠ¸ë¥¼ ì•ˆì „í•˜ê²Œ ì¶”ì¶œ"""
        for selector in selectors:
            try:
                found_element = element.find_element(By.CSS_SELECTOR, selector)
                text = found_element.text.strip()
                if text:
                    return text
            except:
                continue
        return default_value

    def safe_extract_review_content(self, review):
        """ë¦¬ë·° ë‚´ìš©ì„ ì•ˆì „í•˜ê²Œ ì¶”ì¶œ (HTML êµ¬ì¡°ì— ë§ëŠ” ì…€ë ‰í„° ìš°ì„ )"""
        content_selectors = [
            'div.KqJ8Qqw082 span.MX91DFZo2F',  # HTMLì—ì„œ í™•ì¸ëœ ì •í™•í•œ ê²½ë¡œ
            'div.AlfkEF45qI div.KqJ8Qqw082 span.MX91DFZo2F',  # ë” êµ¬ì²´ì ì¸ ê²½ë¡œ
            'span.MX91DFZo2F',
            'div[class*="content"] span',
            '.review-content',
            'p[class*="content"]',
            'div[class*="text"]'
        ]
        
        for selector in content_selectors:
            try:
                content_spans = review.find_elements(By.CSS_SELECTOR, selector)
                if content_spans:
                    # ê°€ì¥ ê¸´ í…ìŠ¤íŠ¸ë¥¼ ë¦¬ë·° ë‚´ìš©ìœ¼ë¡œ ì„ íƒ
                    contents = [span.text.strip() for span in content_spans if span.text.strip()]
                    if contents:
                        # ë¦¬ë·° ë‚´ìš©ì€ ë³´í†µ ê°€ì¥ ê¸´ í…ìŠ¤íŠ¸ì´ë¯€ë¡œ ê¸¸ì´ë¡œ íŒë‹¨
                        longest_content = max(contents, key=len)
                        # ë„ˆë¬´ ì§§ì€ í…ìŠ¤íŠ¸ëŠ” ì œì™¸ (ë‚ ì§œë‚˜ ì´ë¦„ ë“±)
                        if len(longest_content) > 10:
                            return longest_content
            except:
                continue
        return ""

    def determine_review_type(self, review):
        """ë¦¬ë·° ìœ í˜•ì„ ê²°ì • (HTML êµ¬ì¡°ì— ë§ëŠ” ì…€ë ‰í„° ìš°ì„ )"""
        review_type = "ì¼ë°˜ë¦¬ë·°"
        try:
            type_selectors = [
                'span.W1IZsaUmnu',  # HTMLì—ì„œ í™•ì¸ëœ ì •í™•í•œ í´ë˜ìŠ¤
                'div.KqJ8Qqw082 span.W1IZsaUmnu',  # ë” êµ¬ì²´ì ì¸ ê²½ë¡œ
                'div.AlfkEF45qI div.KqJ8Qqw082 span.W1IZsaUmnu',
                'div.KqJ8Qqw082 span',
                'div.AlfkEF45qI div.KqJ8Qqw082 span',
                'span[class*="tag"]',
                '.review-tag',
                'div[class*="type"]'
            ]
            
            for selector in type_selectors:
                try:
                    type_elements = review.find_elements(By.CSS_SELECTOR, selector)
                    types = [elem.text.strip() for elem in type_elements if elem.text.strip()]
                    
                    # í•œë‹¬ì‚¬ìš©, ì¬êµ¬ë§¤ ë“± í™•ì¸
                    is_one_month_review = any("í•œë‹¬ì‚¬ìš©" in t for t in types)
                    is_reorder = any("ì¬êµ¬ë§¤" in t for t in types)
                    
                    if is_one_month_review and is_reorder:
                        review_type = "í•œë‹¬+ì¬êµ¬ë§¤"
                    elif is_one_month_review:
                        review_type = "í•œë‹¬ì‚¬ìš©ê¸°"
                    elif is_reorder:
                        review_type = "ì¬êµ¬ë§¤"
                    break
                except:
                    continue
        except Exception as e:
            self.update_log(f"ë¦¬ë·° ìœ í˜• í™•ì¸ ì‹¤íŒ¨: {e}")
        return review_type

    def extract_review_photos(self, review, photo_folder_path, current_page, review_index):
        """ë¦¬ë·° ì´ë¯¸ì§€ë¥¼ ì•ˆì „í•˜ê²Œ ì¶”ì¶œ"""
        self.update_log(f"      ğŸ“¸ ë¦¬ë·° ì´ë¯¸ì§€ ì¶”ì¶œ ì‹œì‘...")
        photos = []
        try:
            photo_selectors = [
                'img.UpImHAUeYJ[alt="review_image"]',  # HTMLì—ì„œ í™•ì¸ëœ ì •í™•í•œ í´ë˜ìŠ¤
                'div.AlfkEF45qI img.UpImHAUeYJ[alt="review_image"]',  # ë” êµ¬ì²´ì ì¸ ê²½ë¡œ
                'div.s30AvhHfb0 img.UpImHAUeYJ[alt="review_image"]',  # ì´ë¯¸ì§€ ì»¨í…Œì´ë„ˆ ë‚´ë¶€
                'img[alt="review_image"]',
                'img[class*="review"]',
                '.review-photo img'
            ]
            
            self.update_log(f"      ğŸ” ì´ë¯¸ì§€ ìš”ì†Œë¥¼ ì°¾ëŠ” ì¤‘...")
            photo_elements = []
            for selector in photo_selectors:
                try:
                    elements = review.find_elements(By.CSS_SELECTOR, selector)
                    if elements:
                        photo_elements = elements
                        self.update_log(f"      âœ… ì´ë¯¸ì§€ ìš”ì†Œ {len(elements)}ê°œ ë°œê²¬: {selector}")
                        break
                except:
                    continue
            
            if not photo_elements:
                self.update_log(f"      âš ï¸ ì´ë¯¸ì§€ ìš”ì†Œë¥¼ ì°¾ì„ ìˆ˜ ì—†ìŠµë‹ˆë‹¤.")
                return photos
            
            for i, photo_element in enumerate(photo_elements):
                self.update_log(f"        ğŸ–¼ï¸ ì´ë¯¸ì§€ {i+1}/{len(photo_elements)} ì²˜ë¦¬ ì¤‘...")
                try:
                    # data-src ì†ì„±ì„ ìš°ì„ ì ìœ¼ë¡œ ì‚¬ìš© (ì›ë³¸ ì´ë¯¸ì§€)
                    self.update_log(f"          ğŸ”— ì´ë¯¸ì§€ URL ì¶”ì¶œ ì¤‘...")
                    photo_url = photo_element.get_attribute('data-src')
                    if not photo_url:
                        photo_url = photo_element.get_attribute('src')
                    
                    if not photo_url:
                        self.update_log(f"          âš ï¸ ì´ë¯¸ì§€ URLì„ ì°¾ì„ ìˆ˜ ì—†ìŠµë‹ˆë‹¤.")
                        continue
                    
                    self.update_log(f"          ğŸ“ ì›ë³¸ URL: {photo_url}")
                    
                    # ë„¤ì´ë²„ ì´ë¯¸ì§€ URLì—ì„œ íŒŒë¼ë¯¸í„° ì œê±°í•˜ì—¬ ì›ë³¸ ê°€ì ¸ì˜¤ê¸°
                    if 'pstatic.net' in photo_url:
                        original_url = photo_url.split('?')[0]
                    else:
                        original_url = photo_url
                    
                    self.update_log(f"          ğŸ”„ ì²˜ë¦¬ëœ URL: {original_url}")
                    
                    # ì´ë¯¸ì§€ ë‹¤ìš´ë¡œë“œ
                    self.update_log(f"          â¬‡ï¸ ì´ë¯¸ì§€ ë‹¤ìš´ë¡œë“œ ì¤‘...")
                    response = requests.get(original_url, timeout=10)
                    response.raise_for_status()
                    
                    photo_path = os.path.join(photo_folder_path, f'review_page{current_page}_{review_index+1}_photo_{i+1}.jpg')
                    with open(photo_path, 'wb') as file:
                        file.write(response.content)
                    
                    self.update_log(f"          ğŸ’¾ ì´ë¯¸ì§€ ì €ì¥ ì™„ë£Œ: {photo_path}")
                    
                    # ì´ë¯¸ì§€ í¬ê¸° í™•ì¸
                    try:
                        with PILImage.open(photo_path) as img:
                            width, height = img.size
                            self.update_log(f"          âœ… ì´ë¯¸ì§€ ì €ì¥ ì™„ë£Œ: {width}x{height}")
                    except Exception as e:
                        self.update_log(f"          âš ï¸ ì´ë¯¸ì§€ í¬ê¸° í™•ì¸ ì‹¤íŒ¨: {e}")
                    
                    photos.append(photo_path)
                    self.update_log(f"          âœ… ì´ë¯¸ì§€ {i+1} ì²˜ë¦¬ ì™„ë£Œ")
                    
                except Exception as e:
                    self.update_log(f"          âŒ ì´ë¯¸ì§€ {i+1} ë‹¤ìš´ë¡œë“œ ì‹¤íŒ¨: {e}")
                    continue
                    
        except Exception as e:
            self.update_log(f"      âŒ ì´ë¯¸ì§€ ì¶”ì¶œ ì‹¤íŒ¨: {e}")
        
        self.update_log(f"      ğŸ“Š ì´ {len(photos)}ê°œ ì´ë¯¸ì§€ ì¶”ì¶œ ì™„ë£Œ")
        return photos

    def save_to_excel(self, reviews_data, excel_path, start_page):
        # ê¸°ì¡´ ë°ì´í„° ì½ê¸° (ì¤‘ë³µ ì²´í¬ìš©)
        existing_reviews = set()
        if os.path.exists(excel_path):
            wb = load_workbook(excel_path)
            ws = wb.active
            start_row = ws.max_row + 1
            
            # ê¸°ì¡´ ë¦¬ë·° ë°ì´í„°ë¥¼ setìœ¼ë¡œ ì €ì¥ (ì¤‘ë³µ ì²´í¬ìš©)
            for row in range(2, ws.max_row + 1):  # í—¤ë” ì œì™¸
                reviewer = ws.cell(row=row, column=3).value or ""
                date = ws.cell(row=row, column=4).value or ""
                score = ws.cell(row=row, column=2).value or ""
                content = ws.cell(row=row, column=7).value or ""
                # ì¤‘ë³µ ì²´í¬ìš© í‚¤ ìƒì„± (ë‚´ìš© ì „ì²´)
                review_key = f"{reviewer}_{date}_{score}_{content}"
                existing_reviews.add(review_key)
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Reviews"
            headers = ["Page", "Review Score", "Reviewer Name", "Review Date", "Product(Option) Name", "Review Type", "Content", "Photo"]
            ws.append(headers)
            for i, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=i, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            start_row = 2

        # ì»¬ëŸ¼ ë„ˆë¹„ ì„¤ì •
        column_widths = {
            "A": 10, "B": 12, "C": 15, "D": 15, "E": 25, "F": 15, "G": 60, "H": 12
        }
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        # ì¤‘ë³µë˜ì§€ ì•Šì€ ë¦¬ë·°ë§Œ í•„í„°ë§
        new_reviews = []
        duplicate_count = 0
        
        for review in reviews_data:
            # ì¤‘ë³µ ì²´í¬ìš© í‚¤ ìƒì„± (ë‚´ìš© ì „ì²´)
            reviewer = review["Reviewer Name"] or ""
            date = review["Review Date"] or ""
            score = review["Review Score"] or ""
            content = review["Content"] or ""
            review_key = f"{reviewer}_{date}_{score}_{content}"
            
            if review_key not in existing_reviews:
                new_reviews.append(review)
                existing_reviews.add(review_key)  # ë‹¤ìŒ ì¤‘ë³µ ì²´í¬ë¥¼ ìœ„í•´ ì¶”ê°€
            else:
                duplicate_count += 1
        
        if duplicate_count > 0:
            self.update_log(f"ì¤‘ë³µ ë¦¬ë·° {duplicate_count}ê°œë¥¼ ì œì™¸í–ˆìŠµë‹ˆë‹¤.")

        # ë¦¬ë·° ë°ì´í„° ì¶”ê°€
        for row_index, review in enumerate(new_reviews, start=start_row):
            ws.cell(row=row_index, column=1, value=start_page).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=row_index, column=2, value=review["Review Score"]).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.cell(row=row_index, column=3, value=review["Reviewer Name"]).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.cell(row=row_index, column=4, value=review["Review Date"]).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.cell(row=row_index, column=5, value=review["Product(Option) Name"]).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.cell(row=row_index, column=6, value=review["Review Type"]).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.cell(row=row_index, column=7, value=review["Content"]).alignment = Alignment(vertical="center", wrap_text=True)

            start_page += 1

        # í–‰ ë†’ì´ ì„¤ì •
        for row in range(start_row, ws.max_row + 1):
            ws.row_dimensions[row].height = 75

        wb.save(excel_path)
        self.update_log(f"ìƒˆë¡œìš´ ë¦¬ë·° {len(new_reviews)}ê°œë¥¼ {excel_path}ì— ì €ì¥í–ˆìŠµë‹ˆë‹¤.")

class ReviewCrawlerGUI:
    def __init__(self, root):
        self.root = root
        self.root.title('ë¦¬ë·° ìˆ˜ì§‘ê¸°')
        self.root.geometry('500x600')
        
        # ë³€ìˆ˜ë“¤
        self.site_var = tk.StringVar(value="naver")
        self.sort_var = tk.StringVar(value="ranking")
        self.page_var = tk.StringVar(value="5")
        self.custom_page_var = tk.StringVar()
        self.open_folder_var = tk.BooleanVar()
        self.search_keyword_var = tk.StringVar(value="ì˜¤í˜ë¼ê¸€ë¼ìŠ¤")
        
        self.initUI()

    def initUI(self):
        # ë©”ì¸ í”„ë ˆì„
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # ê²€ìƒ‰ì–´ ì…ë ¥
        ttk.Label(main_frame, text="ê²€ìƒ‰ì–´:").grid(row=0, column=0, sticky=tk.W, pady=5)
        self.search_input = ttk.Entry(main_frame, textvariable=self.search_keyword_var, width=50)
        self.search_input.grid(row=0, column=1, columnspan=3, sticky=(tk.W, tk.E), pady=5)
        
        # ì‚¬ì´íŠ¸ ì„ íƒ
        ttk.Label(main_frame, text="ì‚¬ì´íŠ¸:").grid(row=1, column=0, sticky=tk.W, pady=5)
        ttk.Radiobutton(main_frame, text="ë„¤ì´ë²„", variable=self.site_var, value="naver").grid(row=1, column=1, pady=5)
        ttk.Radiobutton(main_frame, text="ì¿ íŒ¡", variable=self.site_var, value="coupang").grid(row=1, column=2, pady=5)
        
        # ì •ë ¬ ìˆœì„œ
        ttk.Label(main_frame, text="ì •ë ¬ ìˆœì„œ:").grid(row=2, column=0, sticky=tk.W, pady=5)
        ttk.Radiobutton(main_frame, text="ë­í‚¹ìˆœ", variable=self.sort_var, value="ranking").grid(row=2, column=1, pady=5)
        ttk.Radiobutton(main_frame, text="ìµœì‹ ìˆœ", variable=self.sort_var, value="latest").grid(row=2, column=2, pady=5)
        ttk.Radiobutton(main_frame, text="í‰ì ë‚®ì€ìˆœ", variable=self.sort_var, value="lowest").grid(row=2, column=3, pady=5)
        
        # ë¦¬ë·° í˜ì´ì§€
        ttk.Label(main_frame, text="ë¦¬ë·° í˜ì´ì§€:").grid(row=3, column=0, sticky=tk.W, pady=5)
        ttk.Radiobutton(main_frame, text="5", variable=self.page_var, value="5").grid(row=3, column=1, pady=5)
        ttk.Radiobutton(main_frame, text="15", variable=self.page_var, value="15").grid(row=3, column=2, pady=5)
        ttk.Radiobutton(main_frame, text="50", variable=self.page_var, value="50").grid(row=3, column=3, pady=5)
        ttk.Radiobutton(main_frame, text="max", variable=self.page_var, value="max").grid(row=3, column=4, pady=5)
        
        # ì§ì ‘ ì…ë ¥
        ttk.Label(main_frame, text="ì§ì ‘ ì…ë ¥:").grid(row=4, column=0, sticky=tk.W, pady=5)
        self.custom_page_entry = ttk.Entry(main_frame, textvariable=self.custom_page_var, width=10)
        self.custom_page_entry.grid(row=4, column=1, sticky=tk.W, pady=5)
        self.custom_page_entry.bind('<KeyRelease>', self.on_custom_page_change)
        
        # ì €ì¥ ê²½ë¡œ
        ttk.Label(main_frame, text="ì €ì¥ ê²½ë¡œ:").grid(row=5, column=0, sticky=tk.W, pady=5)
        self.path_input = ttk.Entry(main_frame, width=40)
        self.path_input.grid(row=5, column=1, columnspan=2, sticky=(tk.W, tk.E), pady=5)
        ttk.Button(main_frame, text="ì„ íƒ", command=self.select_path).grid(row=5, column=3, pady=5)
        
        # í´ë” ìë™ ì—´ê¸° ì˜µì…˜
        ttk.Checkbutton(main_frame, text="í¬ë¡¤ë§ í›„ í´ë” ì—´ê¸°", variable=self.open_folder_var).grid(row=6, column=0, columnspan=4, sticky=tk.W, pady=5)
        
        # ì‹¤í–‰ ë²„íŠ¼
        self.run_button = ttk.Button(main_frame, text="ì‹¤í–‰", command=self.run_crawler)
        self.run_button.grid(row=7, column=0, columnspan=4, pady=10)
        
        # ë¡œê·¸ ì¶œë ¥
        ttk.Label(main_frame, text="ë¡œê·¸:").grid(row=8, column=0, sticky=tk.W, pady=5)
        self.log_text = scrolledtext.ScrolledText(main_frame, height=15, width=60)
        self.log_text.grid(row=9, column=0, columnspan=4, sticky=(tk.W, tk.E, tk.N, tk.S), pady=5)
        
        # ê·¸ë¦¬ë“œ ê°€ì¤‘ì¹˜ ì„¤ì •
        main_frame.columnconfigure(1, weight=1)
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)

    def on_custom_page_change(self, event):
        if self.custom_page_var.get():
            self.page_var.set("custom")

    def select_path(self):
        folder_path = filedialog.askdirectory(title="ì €ì¥ ê²½ë¡œ ì„ íƒ")
        if folder_path:
            self.path_input.delete(0, tk.END)
            self.path_input.insert(0, folder_path)

    def update_log(self, message):
        self.log_text.insert(tk.END, message + "\n")
        self.log_text.see(tk.END)
        self.root.update_idletasks()

    def run_crawler(self):
        # ì…ë ¥ê°’ ê°€ì ¸ì˜¤ê¸°
        search_keyword = self.search_keyword_var.get()
        if not search_keyword:
            messagebox.showerror("ì˜¤ë¥˜", "ê²€ìƒ‰ì–´ë¥¼ ì…ë ¥í•´ì£¼ì„¸ìš”.")
            return
            
        sort_option = self.sort_var.get()
        
        if self.page_var.get() == "custom":
            input_pages = self.custom_page_var.get()
            if not input_pages:
                messagebox.showerror("ì˜¤ë¥˜", "í˜ì´ì§€ ìˆ˜ë¥¼ ì…ë ¥í•´ì£¼ì„¸ìš”.")
                return
        else:
            input_pages = self.page_var.get()

        photo_folder_path = self.path_input.get()
        if not photo_folder_path:
            messagebox.showerror("ì˜¤ë¥˜", "ì €ì¥ ê²½ë¡œë¥¼ ì„ íƒí•´ì£¼ì„¸ìš”.")
            return

        # í¬ë¡¤ë§ ìŠ¤ë ˆë“œ ì‹œì‘
        self.crawler_thread = CrawlerThread(search_keyword, input_pages, photo_folder_path, sort_option, callback=self.update_log)
        self.crawler_thread.daemon = True
        self.crawler_thread.start()

        # UI ë¹„í™œì„±í™”
        self.run_button.config(state='disabled')
        self.update_log("í¬ë¡¤ë§ì„ ì‹œì‘í•©ë‹ˆë‹¤...")
        
        # ì™„ë£Œ ì²´í¬ë¥¼ ìœ„í•œ íƒ€ì´ë¨¸
        self.check_thread_completion()

    def check_thread_completion(self):
        if hasattr(self, 'crawler_thread') and self.crawler_thread.is_alive():
            self.root.after(1000, self.check_thread_completion)
        else:
            self.run_button.config(state='normal')
            self.update_log("í¬ë¡¤ë§ì´ ì™„ë£Œë˜ì—ˆìŠµë‹ˆë‹¤.")
            
            # í´ë” ìë™ ì—´ê¸°
            if self.open_folder_var.get():
                folder_path = self.path_input.get()
                if os.path.exists(folder_path):
                    if sys.platform == 'win32':
                        os.startfile(folder_path)
                    elif sys.platform == 'darwin':  # macOS
                        subprocess.Popen(['open', folder_path])
                    else:  # linux
                        subprocess.Popen(['xdg-open', folder_path])

if __name__ == '__main__':
    root = tk.Tk()
    app = ReviewCrawlerGUI(root)
    root.mainloop()
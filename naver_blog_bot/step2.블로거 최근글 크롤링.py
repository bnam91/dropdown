'''
# STEP2.블로거 최근글 크롤링 ver.1.2
- step 1단계에서 추출한 블로거들의 최근 글을 크롤링합니다.

### 방법
- step 1단계에서 추출한 엑셀파일을 그대로 넣고 실행하면 됨
- (a2)는 연결하지 말것. 점령할 키워드는 한번에 하나씩진행할 것 (셀이 안맞아서 C레벨에서 이상하게 나옴)
- 처음 로딩시간이 제법 걸리는 점 참고. 3분 이내로 되긴 함

'''


import tkinter as tk
from tkinter import filedialog
import re
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import os
from datetime import datetime

def select_excel_file():
    root = tk.Tk()
    root.withdraw()  # 루트 창 숨기기
    root.update()  # 상태 갱신

def select_excel_file():
    root = tk.Tk()
    root.withdraw()
    file_path = filedialog.askopenfilename(title="엑셀 파일 선택", filetypes=[("Excel files", "*.xlsx")])
    return file_path

def setup_webdriver():
    options = Options()
    options.headless = False
    driver = webdriver.Chrome(options=options)
    return driver


def extract_blog_ids(file_path):
    wb = load_workbook(file_path)
    ws = wb.active
    urls, names = [], []
    for row in range(1, ws.max_row + 1):
        url = ws['E' + str(row)].value
        name = ws['A' + str(row)].value
        if url and url.startswith("https://blog.naver.com/"):
            blog_id = url.split('/')[3]
            urls.append(f"https://blog.naver.com/PostList.naver?blogId={blog_id}&skinType=&skinId=&from=menu")
            names.append(name)
        time.sleep(1)
    return urls, names



def set_15_line_view(driver):
    select_box = driver.find_elements(By.CSS_SELECTOR, 'a.btn_select.pcol2._ListCountToggle._returnFalse')
    #목록열려있을 때 --> 만약, 셀렉박스가 보이면 클릭해서 15열이 보이게 만들어라. -->
    if select_box:
        select_box[0].click()
        time.sleep(1)
        fifteen_lines_option = WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable((By.XPATH, "//a[@data-value='15']"))
        )
        fifteen_lines_option.click()
        time.sleep(2)

    #목록 닫혀있다면 --> 목록열기 먼저 누를 것.
    else:
        open_list_button = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.CLASS_NAME, 'btn_openlist.pcol2._toggleTopList._returnFalse'))
        )
        open_list_button.click()
        time.sleep(1)
        
        # 그다음 셀렉박스 열어서 15개로 맞출 것
        select_button = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, 'a.btn_select.pcol2._ListCountToggle._returnFalse'))
        )
        select_button.click()
        time.sleep(1)
        fifteen_lines_option = WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable((By.XPATH, "//a[@data-value='15']"))
        )
        fifteen_lines_option.click()
        time.sleep(2)


def scrape_blog_data(driver, url):
    driver.get(url)
    time.sleep(5)
    set_15_line_view(driver)
    try:
        WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, 'table.blog2_list'))
        )
        posts = driver.find_elements(By.CSS_SELECTOR, 'table.blog2_list tbody tr')[:15]
        data = []
        for post in posts:
            title_elements = post.find_elements(By.CSS_SELECTOR, 'span.ell2.pcol2')
            date_elements = post.find_elements(By.CSS_SELECTOR, 'div.wrap_td span.date.pcol2')
            if title_elements and date_elements:
                title = title_elements[0].text
                date = date_elements[0].text
                data.append((title, date))
    except Exception as e:
        print(f"Error processing {url}: {str(e)}")
        data = []
    return data



def apply_excel_styles(ws):
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="FCD5B4", end_color="FCD5B4", fill_type="solid")
    center_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    grey_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
    
    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 70
    ws.column_dimensions['D'].width = 16

    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = center_alignment
        cell.fill = header_fill
        cell.border = thin_border
    
    return grey_fill

def save_data_to_excel(urls, names, data_list, ws, grey_fill):
    for index, (url, name, data) in enumerate(zip(urls, names, data_list)):
        for title, date in data:
            row = [name, "", title, date]
            ws.append(row)
            if index % 2 == 1:
                for cell in ws[ws.max_row]:
                    cell.fill = grey_fill

def main():
    file_path = select_excel_file()
    if not file_path:
        return

    # 파일명에서 키워드 추출
    input_filename = os.path.basename(file_path)
    keyword = input_filename.split('_')[1]  # '저당잼' 추출
    
    # 오늘 날짜 생성
    today = datetime.now().strftime('%Y%m%d')
    
    # 새 파일명 생성
    output_filename = f'step02_{keyword}_블로거최근글_{today}.xlsx'
    output_path = os.path.join(os.path.dirname(file_path), output_filename)

    driver = setup_webdriver()
    urls, names = extract_blog_ids(file_path)
    data_list = [scrape_blog_data(driver, url) for url in urls]

    wb = Workbook()
    ws = wb.active
    ws.append(["NAME", "2차 키워드", "글 제목", "작성일"])
    grey_fill = apply_excel_styles(ws)
    save_data_to_excel(urls, names, data_list, ws, grey_fill)

    wb.save(output_path)
    print(f"파일이 저장되었습니다: {output_path}")

    driver.quit()

if __name__ == "__main__":
    main()

    ## 이게 거의 최종일듯

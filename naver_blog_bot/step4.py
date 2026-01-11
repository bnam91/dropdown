'''
스텝3은 웹에서 2차키워드를 추출하는 것임
스텝4는 추출한 2차키워드에서 criteria 찾는 것


'''

import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from bs4 import BeautifulSoup
import time
from tkinter import Tk, filedialog
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

def get_excel_file_path():
    root = Tk()
    root.withdraw()
    file_path = filedialog.askopenfilename(title='DB 엑셀 파일 선택', filetypes=[("Excel files", "*.xlsx *.xls")])
    return file_path

def load_keywords_from_excel(file_path):
    df = pd.read_excel(file_path, header=0)
    print(df.columns)
    if df.columns[0] not in df or df.columns[1] not in df:
        print("필요한 열이 엑셀 파일에 없습니다.")
        return pd.DataFrame()
    return df.dropna(subset=[df.columns[1]])

def save_results_to_excel(data_dict, keyword_index):
    output_path = filedialog.asksaveasfilename(defaultextension='.xlsx', title='저장할 파일 경로를 지정하세요')
    if not output_path:
        print("저장 경로가 지정되지 않았습니다.")
        return
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        for sheet_name, records in data_dict.items():
            df = pd.DataFrame(records, columns=['2차키워드', '검색량', 'NAME', '제목', '게시일', 'URL', '이메일주소', '상위노출여부'])
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            worksheet = writer.sheets[sheet_name]

            # 열 너비 설정
            column_widths = {'A': 16, 'B': 16, 'C': 16, 'D': 70, 'E': 16, 'F': 8, 'G': 24, 'H': 16}
            for col, width in column_widths.items():
                worksheet.column_dimensions[col].width = width

            # 스타일 설정
            red_bold_font = Font(bold=True, color="FF0000")
            grey_fill = PatternFill(start_color='EAEAEA', end_color='EAEAEA', fill_type="solid")
            standard_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                     top=Side(style='thin'), bottom=Side(style='thin'))

            for row in range(2, worksheet.max_row + 1):
                for col in range(1, 9):  # A to H
                    cell = worksheet.cell(row=row, column=col)
                    cell.border = standard_border
                    
                    # 짝수번째 키워드 색상 적용
                    if keyword_index[df.iloc[row-2, 0]] % 2 == 1:
                        cell.fill = grey_fill
                    
                    # C열에 시트 제목과 동일한 내용이 있다면
                    if worksheet.cell(row=row, column=3).value == sheet_name:
                        cell.font = red_bold_font
                        if col == 8:  # H열에 "criteria" 입력
                            worksheet.cell(row=row, column=8).value = "criteria"
                
            # 헤더 스타일
            for col in range(1, 9):
                cell = worksheet.cell(row=1, column=col)
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color='CCC0DA', end_color='CCC0DA', fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                     top=Side(style='thin'), bottom=Side(style='thin'))

            # 필터 추가
            worksheet.auto_filter.ref = "A1:H" + str(worksheet.max_row)

    print("데이터가 성공적으로 저장되었습니다.")


options = Options()
options.headless = False

driver = webdriver.Chrome(options=options)

file_path = get_excel_file_path()
if not file_path:
    driver.quit()
    exit()

df_keywords = load_keywords_from_excel(file_path)
if df_keywords.empty:
    driver.quit()
    exit()

data_dict = {}
keyword_index = {}

try:
    for index, row in df_keywords.iterrows():
        group_name, keyword = row[df_keywords.columns[0]], row[df_keywords.columns[1]]
        if keyword not in keyword_index:
            keyword_index[keyword] = len(keyword_index)
        driver.get('https://www.naver.com')
        time.sleep(1)

        driver.find_element(By.ID, "query").send_keys(keyword)
        driver.find_element(By.ID, "search-btn").click()
        time.sleep(1)

        driver.find_element(By.LINK_TEXT, '블로그').click()
        time.sleep(2)

        driver.execute_script('window.scrollTo(0, document.body.scrollHeight);')
        time.sleep(2)

        html = BeautifulSoup(driver.page_source, 'html.parser')
        name_links = html.find_all('a', class_='name')[:15]
        title_links = html.find_all('a', class_='title_link')[:15]
        date_links = html.find_all('span', class_='sub')[:15]

        data = []
        for name_link, title_link, date_link in zip(name_links, title_links, date_links):
            url = title_link['href']
            email = ""
            if "https://blog.naver.com/" in url:
                username = url.split('/')[3]
                email = f"{username}@naver.com"
            elif "https://adcr.naver.com/" in url:
                email = "파워콘텐츠"
            elif "https://post.naver.com/" in url:
                email = "포스트"

            data.append([keyword, '', name_link.text.strip(), title_link.text.strip(), date_link.text.strip(), url, email, ''])

        if group_name in data_dict:
            data_dict[group_name].extend(data)
        else:
            data_dict[group_name] = data
finally:
    driver.quit()

if data_dict:
    save_results_to_excel(data_dict, keyword_index)
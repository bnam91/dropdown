'''
# STEP 1-1. 단일 키워드 
- 현재 검색어 상위노출 중인 블로거를 추출합니다. (단일 키워드)

'''

import tkinter as tk
from tkinter import messagebox
from tkinter import simpledialog
import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from bs4 import BeautifulSoup
import time
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import os

# Tkinter로 입력 폼 구현
class InputDialog(simpledialog.Dialog):
    def body(self, master):
        self.title('네이버 블로그 검색')
        tk.Label(master, text='검색어를 입력하세요:').grid(row=0, column=0, sticky='e')
        self.search_entry = tk.Entry(master, width=35)
        self.search_entry.grid(row=0, column=1, columnspan=4)

        tk.Label(master, text='가져올 글의 수:').grid(row=1, column=0, sticky='e')
        self.num_var = tk.StringVar(value='15')
        self.radio_15 = tk.Radiobutton(master, text='15개', variable=self.num_var, value='15', command=self.clear_custom)
        self.radio_30 = tk.Radiobutton(master, text='30개', variable=self.num_var, value='30', command=self.clear_custom)
        self.radio_50 = tk.Radiobutton(master, text='50개', variable=self.num_var, value='50', command=self.clear_custom)
        self.radio_100 = tk.Radiobutton(master, text='100개', variable=self.num_var, value='100', command=self.clear_custom)
        self.radio_15.grid(row=1, column=1)
        self.radio_30.grid(row=1, column=2)
        self.radio_50.grid(row=1, column=3)
        self.radio_100.grid(row=1, column=4)

        tk.Label(master, text='또는 직접 입력:').grid(row=2, column=2, sticky='e')
        self.custom_entry = tk.Entry(master, width=10)
        self.custom_entry.grid(row=2, column=3, columnspan=2)
        self.custom_entry.bind('<KeyRelease>', self.clear_radio)
        return self.search_entry

    def clear_custom(self):
        self.custom_entry.delete(0, tk.END)

    def clear_radio(self, event):
        if self.custom_entry.get():
            self.num_var.set('')

    def validate(self):
        self.errors = []
        self.search_query = self.search_entry.get().strip()
        self.custom_num = self.custom_entry.get().strip()
        self.selected_num = self.num_var.get()
        if not self.search_query:
            self.errors.append('검색어를 입력해야 합니다.')
        if self.custom_num and self.selected_num:
            self.errors.append('직접 입력과 선택 옵션 중 하나만 사용해주세요.')
        if self.custom_num:
            try:
                self.num_of_posts = int(self.custom_num)
            except ValueError:
                self.errors.append('직접 입력한 수는 유효한 숫자여야 합니다.')
        elif self.selected_num:
            self.num_of_posts = int(self.selected_num)
        else:
            self.errors.append('글 수를 입력하거나 선택해주세요.')
        if self.errors:
            messagebox.showerror('입력 오류', '\n'.join(self.errors))
            return False
        return True

    def apply(self):
        self.result = (self.search_query, self.num_of_posts)

# Tkinter로 입력 받기
root = tk.Tk()
root.withdraw()
dialog = InputDialog(root)
if dialog.result is None:
    raise SystemExit
search_query, num_of_posts = dialog.result

options = Options()
options.headless = False
driver = webdriver.Chrome(options=options)

driver.get("https://www.naver.com")
time.sleep(1)

driver.find_element(By.ID, "query").send_keys(search_query)
driver.find_element(By.ID, "search-btn").click()
time.sleep(1)

driver.find_element(By.LINK_TEXT, "블로그").click()
time.sleep(2)

try:
    while True:
        driver.execute_script('window.scrollTo(0, document.body.scrollHeight);')
        time.sleep(2)
        html = BeautifulSoup(driver.page_source, 'html.parser')
        posts = html.find_all('a', class_='title_link')
        if (num_of_posts and len(posts) >= num_of_posts) or html.find('span', class_='bl_tit'):
            break

    if num_of_posts is None or len(posts) < num_of_posts:
        num_of_posts = len(posts)

except Exception as e:
    messagebox.showerror('브라우저 창이 닫혔습니다:', str(e))
    driver.quit()
    raise SystemExit

name_links = html.find_all('a', class_='name')[:num_of_posts]
title_links = posts[:num_of_posts]
date_links = html.find_all('span', class_='sub')[:num_of_posts]

data = []
for name_link, title_link, date_link in zip(name_links, title_links, date_links):
    url = title_link['href']
    email = ''
    if "https://blog.naver.com/" in url:
        username = url.split('/')[3]
        email = f"{username}@naver.com"
    elif "https://adcr.naver.com/" in url:
        email = "파워콘텐츠"
    elif "https://post.naver.com/" in url:
        email = "포스트"

    data.append((name_link.text.strip(), "", title_link.text.strip(), date_link.text.strip(), url, email))

df = pd.DataFrame(data, columns=['NAME', '2차키워드', '제목', '게시일', 'URL', '이메일'])

wb = Workbook()
ws = wb.active

for r in dataframe_to_rows(df, index=False, header=True):
    ws.append(r)

ws.column_dimensions['A'].width = 16
ws.column_dimensions['B'].width = 16
ws.column_dimensions['C'].width = 70
ws.column_dimensions['D'].width = 16
ws.column_dimensions['E'].width = 8
ws.column_dimensions['F'].width = 24

center_alignment = Alignment(horizontal='center')
for cell in ws["D"] + ws["F"]:
    cell.alignment = center_alignment

header_font = Font(bold=True)
header_fill = PatternFill(start_color="B7DEE8", end_color="B7DEE8", fill_type="solid")
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

for cell in ws[1]:
    cell.font = header_font
    cell.alignment = center_alignment
    cell.fill = header_fill
    cell.border = thin_border

gray_fill = PatternFill(start_color="ECECEC", end_color="ECECEC", fill_type="solid")
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    if row[5].value in ['파워콘텐츠', '포스트']:
        for cell in row:
            cell.fill = gray_fill

ws.auto_filter.ref = ws.dimensions

# 현재 날짜를 포함한 폴더명 생성
current_date = datetime.now().strftime('%Y%m%d')
folder_name = f'results_{search_query}_{current_date}'

# 폴더 생성
if not os.path.exists(folder_name):
    os.makedirs(folder_name)

# 오늘 날짜를 YYYYMMDD 형식으로 가져오기
today = datetime.now().strftime('%Y%m%d')
# 파일명 생성
save_filename = f'{folder_name}/step01_{search_query}_{num_of_posts}_{today}.xlsx'

# 파일 저장
wb.save(save_filename)
print(f'{save_filename} 파일이 저장되었습니다.')

driver.quit()

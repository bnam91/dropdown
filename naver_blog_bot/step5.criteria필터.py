"""
이 스크립트는 엑셀 파일에서 특정 조건에 따라 데이터를 필터링하는 프로그램입니다.

주요 프로세스:
1. 사용자로부터 엑셀 파일을 선택받습니다.
2. 선택된 파일을 복사하여 새로운 파일을 생성합니다.
3. 엑셀 파일의 모든 시트에서 그룹별로 데이터를 분석합니다.
4. 각 그룹에서 다음 조건에 따라 데이터를 필터링합니다:
   - 'criteria' 텍스트가 있는 행을 찾습니다.
   - '포스트' 또는 '파워콘텐츠'가 있는 행을 삭제합니다.
   - 'criteria'가 발견된 경우, 해당 행부터 그룹 끝까지의 모든 행을 삭제합니다.
   - 'criteria'가 없는 그룹은 전체 삭제합니다.
5. 필터링된 결과를 새로운 파일로 저장합니다.
6. 모든 시트의 데이터를 하나의 시트로 통합합니다.

주요 로직:
- 그룹은 A열의 값이 변경되는 부분을 기준으로 구분됩니다.
- 각 그룹은 시작 행과 끝 행으로 구성됩니다.
- 8번째 칼럼에서 'criteria' 텍스트를 검사합니다.
- 7번째 칼럼에서 '포스트' 또는 '파워콘텐츠'를 검사합니다.
- 행 삭제는 역순으로 진행하여 인덱스 오류를 방지합니다.
- 모든 시트의 데이터는 '통합' 시트에 순차적으로 복사됩니다.
"""

import openpyxl
from tkinter import Tk
from tkinter.filedialog import askopenfilename
import os
import shutil

def filter_groups():
    # GUI를 통해 파일 선택
    Tk().withdraw()
    input_file = askopenfilename(title="엑셀 파일 선택", filetypes=[("Excel files", "*.xlsx;*.xlsm")])
    
    if not input_file:
        print("파일이 선택되지 않았습니다.")
        return

    # 파일 복사
    file_dir = os.path.dirname(input_file)
    file_name = os.path.basename(input_file)
    name, ext = os.path.splitext(file_name)
    output_file = os.path.join(file_dir, f"{name}_filtered{ext}")
    
    # 파일 복사
    shutil.copy2(input_file, output_file)
    
    # 복사된 파일 열기
    wb = openpyxl.load_workbook(output_file)
    
    # 모든 시트에 대해 처리
    for sheet_name in wb.sheetnames:
        print(f"\n시트 '{sheet_name}' 처리 중...")
        ws = wb[sheet_name]
        
        # 그룹 이름과 해당 행 번호를 저장할 딕셔너리
        groups = {}
        
        # 데이터의 마지막 행 찾기
        last_row = ws.max_row

        # A열의 그룹 이름을 기반으로 행 번호 수집
        current_group = None
        group_start = 2
        
        for row in range(2, last_row + 1):
            group_name = ws.cell(row=row, column=1).value
            
            # 새로운 그룹을 만나거나 마지막 행인 경우
            if current_group != group_name:
                if current_group is not None:
                    end_row = row - 1
                    if current_group not in groups:
                        groups[current_group] = []
                    groups[current_group].append((group_start, end_row))
                current_group = group_name
                group_start = row
        
        # 마지막 그룹 처리
        if current_group is not None:
            if current_group not in groups:
                groups[current_group] = []
            groups[current_group].append((group_start, last_row))

        # 각 그룹에 대해 조건 확인하고 불필요한 행 삭제
        rows_to_delete = []
        
        for group_name, ranges in groups.items():
            for start_row, end_row in ranges:
                criteria_found = False
                criteria_row = None
                
                # 현재 그룹 범위에서 'criteria' 텍스트와 특정 키워드 찾기
                for row in range(start_row, end_row + 1):
                    # criteria 체크 (8번째 칼럼)
                    cell_criteria = ws.cell(row=row, column=8)
                    if cell_criteria.value == 'criteria':
                        criteria_found = True
                        criteria_row = row
                        print(f"'criteria' 텍스트 발견: 시트 '{sheet_name}', 그룹 {group_name}, 행 {row}")
                        break
                    
                    # 포스트/파워콘텐츠 체크 (7번째 칼럼, 인덱스 6)
                    cell_type = ws.cell(row=row, column=7)
                    if cell_type.value in ['포스트', '파워콘텐츠']:
                        rows_to_delete.append(row)
                        print(f"'{cell_type.value}' 발견: 시트 '{sheet_name}', 그룹 {group_name}, 행 {row}")
                
                # criteria가 발견된 경우 criteria 행부터 그룹 끝까지 삭제 목록에 추가
                if criteria_found:
                    for row in range(criteria_row, end_row + 1):
                        rows_to_delete.append(row)
                else:
                    # criteria가 없는 경우 그룹의 모든 행을 삭제 목록에 추가
                    for row in range(start_row, end_row + 1):
                        rows_to_delete.append(row)
        
        # 삭제할 행을 역순으로 정렬 (위에서부터 삭제하면 인덱스가 바뀌므로)
        for row in sorted(set(rows_to_delete), reverse=True):
            ws.delete_rows(row)
        
        print(f"시트 '{sheet_name}' 처리 완료")

    # 통합 시트 생성
    print("\n통합 시트 생성 중...")
    merged_sheet = wb.create_sheet("통합")
    
    # 헤더 복사 (첫 번째 시트의 헤더를 사용)
    first_sheet = wb[wb.sheetnames[0]]
    for col in range(1, first_sheet.max_column + 1):
        merged_sheet.cell(row=1, column=col, value=first_sheet.cell(row=1, column=col).value)
    
    # 각 시트의 데이터를 통합 시트에 복사
    current_row = 2
    for sheet_name in wb.sheetnames:
        if sheet_name == "통합":
            continue
            
        ws = wb[sheet_name]
        print(f"시트 '{sheet_name}' 데이터 복사 중... (총 {ws.max_row - 1}행)")
        
        # 각 행의 데이터 복사
        for row in range(2, ws.max_row + 1):
            # 행이 비어있는지 확인
            row_empty = True
            for col in range(1, ws.max_column + 1):
                if ws.cell(row=row, column=col).value is not None:
                    row_empty = False
                    break
            
            # 비어있지 않은 행만 복사
            if not row_empty:
                for col in range(1, ws.max_column + 1):
                    cell_value = ws.cell(row=row, column=col).value
                    if cell_value is not None:  # None 값은 복사하지 않음
                        merged_sheet.cell(row=current_row, column=col, value=cell_value)
                current_row += 1
    
    print(f"통합 시트 생성 완료 (총 {current_row - 1}행)")

    # 통합 시트를 첫 번째 시트로 이동
    wb.move_sheet("통합", offset=-wb.index(wb["통합"]))
    print("통합 시트를 첫 번째 위치로 이동했습니다.")

    # 파일 저장
    wb.save(output_file)
    print(f"\n결과 파일이 저장되었습니다: {output_file}")

# 함수 실행
filter_groups()

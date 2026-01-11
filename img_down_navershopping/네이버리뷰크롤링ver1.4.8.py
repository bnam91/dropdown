import sys
import tkinter as tk
from tkinter import ttk, filedialog, scrolledtext, messagebox
import threading
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import random
import os
import requests
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment
import subprocess
from PIL import Image as PILImage
import io

class CrawlerThread(threading.Thread):
    def __init__(self, url, pages, folder_path, sort_option, chunk_size=50, callback=None):
        threading.Thread.__init__(self)
        self.url = url
        self.pages = pages
        self.folder_path = folder_path
        self.sort_option = sort_option
        self.chunk_size = chunk_size
        self.callback = callback

    def run(self):
        self.crawl_reviews()

    def update_log(self, message):
        if self.callback:
            self.callback(message)

    def crawl_reviews(self):
        self.update_log("크롤링을 시작합니다...")

        # Selenium 설정
        options = Options()
        options.add_experimental_option("detach", True)
        options.add_argument("disable-blink-features=AutomationControlled")
        options.add_experimental_option("detach", True)
        options.add_experimental_option("excludeSwitches", ["enable-logging"])

        driver = webdriver.Chrome(options=options)

        # 먼저 네이버로 이동
        driver.get("https://www.naver.com")
        self.update_log("네이버에 접속했습니다.")
        time.sleep(random.uniform(2, 5))

        # 변수 정의
        url = self.url
        last_page_to_crawl = int(self.pages) if self.pages.isdigit() else self.pages
        sleep_time = 2

        # 네이버
        product_review_button_selector = "#_productFloatingTab > div > div._27jmWaPaKy._1dDHKD1iiX > ul > li:nth-child(2)"
        new_review_button_selector = '#REVIEW > div > div._2LvIMaBiIO > div._2LAwVxx1Sd > div._1txuie7UTH > ul > li:nth-child(2)'
        bad_review_button_selector = '#REVIEW > div > div._2LvIMaBiIO > div._2LAwVxx1Sd > div._1txuie7UTH > ul > li:nth-child(4)'
        review_selector = "#REVIEW > div > div._2LvIMaBiIO > div._2g7PKvqCKe > ul > li"

        # 웹 페이지 열기
        driver.get(url)
        time.sleep(sleep_time)

        # 페이지 끝까지 스크롤
        try:
            driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            self.update_log("페이지 하단으로 스크롤했습니다.")
        except Exception as e:
            self.update_log(f"스크롤 실패: {e}")
            self.close_alert_if_present(driver)

        time.sleep(sleep_time)

        # 상품평 버튼 클릭
        try:
            product_review_button = driver.find_element(By.CSS_SELECTOR, product_review_button_selector)
            product_review_button.click()
            self.update_log("상품평 버튼을 클릭했습니다.")
        except Exception as e:
            self.update_log(f"상품평 버튼 클릭 실패: {e}")
            self.close_alert_if_present(driver)

        time.sleep(3)

        # 정렬 옵션에 따른 버튼 클릭
        if self.sort_option == "latest":
            try:
                new_review_button = driver.find_element(By.CSS_SELECTOR, new_review_button_selector)
                new_review_button.click()
                self.update_log("최신순 정렬 버튼을 클릭했습니다.")
            except Exception as e:
                self.update_log(f"최신순 정렬 버튼 클릭 실패: {e}")
        elif self.sort_option == "lowest":
            try:
                bad_review_button = driver.find_element(By.CSS_SELECTOR, bad_review_button_selector)
                bad_review_button.click()
                self.update_log("평점 낮은 순 정렬 버튼을 클릭했습니다.")
            except Exception as e:
                self.update_log(f"평점 낮은 순 정렬 버튼 클릭 실패: {e}")

        time.sleep(2)

        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        self.update_log("페이지 하단으로 스크롤했습니다.")

        # 리뷰 추출
        self.crawl_reviews_until_page(driver, last_page_to_crawl, self.folder_path, review_selector)

        # 드라이버 종료
        driver.quit()

        self.update_log("크롤링이 완료되었습니다.")

    def close_alert_if_present(self, driver):
        try:
            alert = driver.switch_to.alert
            alert.dismiss()
            self.update_log("예기치 않은 알림창을 닫았습니다.")
        except:
            pass

    def crawl_reviews_until_page(self, driver, last_page, photo_folder_path, review_selector):
        current_page = 1
        chunk_count = 1
        while True:
            chunk_reviews = []
            chunk_start_page = current_page
            
            for _ in range(self.chunk_size):
                if (last_page != "max" and current_page > last_page):
                    break
                
                reviews_data = self.extract_reviews(driver, photo_folder_path, current_page, review_selector)
                chunk_reviews.extend(reviews_data)
                
                if not reviews_data:  # 더 이상 리뷰가 없으면 종료
                    break
                
                current_page += 1
                
                try:
                    # 현재 페이지 번호 확인
                    current_page_element = driver.find_element(By.CSS_SELECTOR, 'a.UWN4IvaQza[aria-current="true"]')
                    current_page_number = int(current_page_element.text)
                    
                    # 페이지네이션에서 모든 페이지 번호 가져오기
                    page_elements = driver.find_elements(By.CSS_SELECTOR, 'a.UWN4IvaQza')
                    page_numbers = [int(elem.text) for elem in page_elements if elem.text.isdigit()]
                    
                    # 다음 페이지가 현재 페이지네이션에 있는지 확인
                    if current_page in page_numbers:
                        # 다음 페이지 버튼 클릭
                        next_page_button = driver.find_element(By.CSS_SELECTOR, f'a.UWN4IvaQza[aria-current="false"]')
                        driver.execute_script("arguments[0].click();", next_page_button)
                    else:
                        # '다음' 버튼 클릭
                        next_button = driver.find_element(By.CSS_SELECTOR, 'a.fAUKm1ewwo._2Ar8-aEUTq')
                        driver.execute_script("arguments[0].click();", next_button)
                    
                    self.update_log(f"{current_page} 페이지로 이동했습니다.")
                    time.sleep(random.uniform(1, 3))
                except Exception as e:
                    self.update_log(f"페이지 이동 실패: {e}")
                    break
            
            # 청크 데이터 저장
            if chunk_reviews:
                excel_path = os.path.join(self.folder_path, f'reviews_chunk_{chunk_count}.xlsx')
                self.save_to_excel(chunk_reviews, excel_path, chunk_start_page)
                chunk_count += 1
            
            if (last_page != "max" and current_page > last_page) or not reviews_data:
                break

    def extract_reviews(self, driver, photo_folder_path, current_page, review_selector):
        reviews_data = []
        try:
            reviews = driver.find_elements(By.CSS_SELECTOR, review_selector)
            for index, review in enumerate(reviews):
                review_score = review.find_element(By.CSS_SELECTOR, 'em').text
                reviewer_name = review.find_element(By.CSS_SELECTOR, 'strong').text
                review_date_selector = f"div._3HKlxxt8Ii > div.iWGqB6S4Lq > span"
                review_date = review.find_element(By.CSS_SELECTOR, review_date_selector).text
                product_name = review.find_element(By.CSS_SELECTOR, 'div._2FXNMst_ak').text

                try:
                    content = review.find_element(By.CSS_SELECTOR, 'div._3z6gI4oI6l > div._1kMfD5ErZ6 > span._2L3vDiadT9').text
                except:
                    content = None

                # 리뷰 유형 확인
                review_type = "일반리뷰"
                try:
                    review_type_elements = review.find_elements(By.CSS_SELECTOR, 'div._3z6gI4oI6l > div > span.byXA4FP1Bq')
                    if review_type_elements:
                        types = [element.text for element in review_type_elements]
                        if "한달사용기" in types and "재구매" in types:
                            review_type = "한달+재구매"
                        elif "한달사용기" in types:
                            review_type = "한달사용기"
                        elif "재구매" in types:
                            review_type = "재구매"
                except:
                    pass

                # 리뷰 이미지 추출
                photos = []
                try:
                    # 리뷰 이미지 찾기 (프로필 이미지 제외)
                    photo_elements = review.find_elements(By.CSS_SELECTOR, 'div._3Bbv1ae9fg > span._1DOkWFrX74 > img[alt="review_image"]')
                    for i, photo_element in enumerate(photo_elements):
                        # data-src 속성을 우선적으로 사용 (원본 이미지)
                        photo_url = photo_element.get_attribute('data-src')
                        if not photo_url:
                            # data-src가 없으면 src 사용
                            photo_url = photo_element.get_attribute('src')
                        
                        self.update_log(f"원본 URL: {photo_url}")
                        
                        # 네이버 이미지 URL에서 파라미터 제거하여 원본 가져오기
                        if 'pstatic.net' in photo_url:
                            # 모든 파라미터 제거
                            original_url = photo_url.split('?')[0]
                        else:
                            original_url = photo_url
                        
                        self.update_log(f"변경된 URL: {original_url}")
                        
                        # 이미지 다운로드 및 원본 그대로 저장
                        response = requests.get(original_url)
                        photo_path = os.path.join(photo_folder_path, f'review_page{current_page}_{index+1}_photo_{i+1}.jpg')
                        with open(photo_path, 'wb') as file:
                            file.write(response.content)
                        
                        # PIL로 이미지 크기 확인
                        try:
                            with PILImage.open(photo_path) as img:
                                width, height = img.size
                                self.update_log(f"저장된 이미지 크기: {width}x{height}")
                        except Exception as e:
                            self.update_log(f"이미지 크기 확인 실패: {e}")
                        
                        photos.append(photo_path)
                except Exception as e:
                    self.update_log(f"이미지 추출 실패: {e}")

                reviews_data.append({
                    "Review Score": review_score,
                    "Reviewer Name": reviewer_name,
                    "Review Date": review_date,
                    "Product(Option) Name": product_name,
                    "Review Type": review_type,
                    "Content": content,
                    "Photos": photos
                })
                self.update_log(f"리뷰 추출 완료: 페이지 {current_page}, 리뷰 {index+1}")
        except Exception as e:
            self.update_log(f"리뷰 정보 추출 실패: {e}")
        return reviews_data

    def save_to_excel(self, reviews_data, excel_path, start_page):
        # 기존 데이터 읽기 (중복 체크용)
        existing_reviews = set()
        if os.path.exists(excel_path):
            wb = load_workbook(excel_path)
            ws = wb.active
            start_row = ws.max_row + 1
            
            # 기존 리뷰 데이터를 set으로 저장 (중복 체크용)
            for row in range(2, ws.max_row + 1):  # 헤더 제외
                reviewer = ws.cell(row=row, column=3).value or ""
                date = ws.cell(row=row, column=4).value or ""
                score = ws.cell(row=row, column=2).value or ""
                content = ws.cell(row=row, column=7).value or ""
                # 중복 체크용 키 생성 (내용 전체)
                review_key = f"{reviewer}_{date}_{score}_{content}"
                existing_reviews.add(review_key)
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Reviews"
            headers = ["Page", "Review Score", "Reviewer Name", "Review Date", "Product(Option) Name", "Review Type", "Content", "Photo"]
            ws.append(headers)
            for i, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=i, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            start_row = 2

        # 컬럼 너비 설정
        column_widths = {
            "A": 10, "B": 12, "C": 15, "D": 15, "E": 25, "F": 15, "G": 60, "H": 12
        }
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        # 중복되지 않은 리뷰만 필터링
        new_reviews = []
        duplicate_count = 0
        
        for review in reviews_data:
            # 중복 체크용 키 생성 (내용 전체)
            reviewer = review["Reviewer Name"] or ""
            date = review["Review Date"] or ""
            score = review["Review Score"] or ""
            content = review["Content"] or ""
            review_key = f"{reviewer}_{date}_{score}_{content}"
            
            if review_key not in existing_reviews:
                new_reviews.append(review)
                existing_reviews.add(review_key)  # 다음 중복 체크를 위해 추가
            else:
                duplicate_count += 1
        
        if duplicate_count > 0:
            self.update_log(f"중복 리뷰 {duplicate_count}개를 제외했습니다.")

        # 리뷰 데이터 추가
        for row_index, review in enumerate(new_reviews, start=start_row):
            ws.cell(row=row_index, column=1, value=start_page).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=row_index, column=2, value=review["Review Score"]).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.cell(row=row_index, column=3, value=review["Reviewer Name"]).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.cell(row=row_index, column=4, value=review["Review Date"]).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.cell(row=row_index, column=5, value=review["Product(Option) Name"]).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.cell(row=row_index, column=6, value=review["Review Type"]).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.cell(row=row_index, column=7, value=review["Content"]).alignment = Alignment(vertical="center", wrap_text=True)

            start_page += 1

        # 행 높이 설정
        for row in range(start_row, ws.max_row + 1):
            ws.row_dimensions[row].height = 75

        wb.save(excel_path)
        self.update_log(f"새로운 리뷰 {len(new_reviews)}개를 {excel_path}에 저장했습니다.")

class ReviewCrawlerGUI:
    def __init__(self, root):
        self.root = root
        self.root.title('리뷰 수집기')
        self.root.geometry('500x600')
        
        # 변수들
        self.site_var = tk.StringVar(value="naver")
        self.sort_var = tk.StringVar(value="ranking")
        self.page_var = tk.StringVar(value="5")
        self.custom_page_var = tk.StringVar()
        self.open_folder_var = tk.BooleanVar()
        
        self.initUI()

    def initUI(self):
        # 메인 프레임
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # 상품 링크
        ttk.Label(main_frame, text="상품 링크:").grid(row=0, column=0, sticky=tk.W, pady=5)
        self.url_input = ttk.Entry(main_frame, width=50)
        self.url_input.grid(row=0, column=1, columnspan=3, sticky=(tk.W, tk.E), pady=5)
        
        # 사이트 선택
        ttk.Label(main_frame, text="사이트:").grid(row=1, column=0, sticky=tk.W, pady=5)
        ttk.Radiobutton(main_frame, text="네이버", variable=self.site_var, value="naver").grid(row=1, column=1, pady=5)
        ttk.Radiobutton(main_frame, text="쿠팡", variable=self.site_var, value="coupang").grid(row=1, column=2, pady=5)
        
        # 정렬 순서
        ttk.Label(main_frame, text="정렬 순서:").grid(row=2, column=0, sticky=tk.W, pady=5)
        ttk.Radiobutton(main_frame, text="랭킹순", variable=self.sort_var, value="ranking").grid(row=2, column=1, pady=5)
        ttk.Radiobutton(main_frame, text="최신순", variable=self.sort_var, value="latest").grid(row=2, column=2, pady=5)
        ttk.Radiobutton(main_frame, text="평점낮은순", variable=self.sort_var, value="lowest").grid(row=2, column=3, pady=5)
        
        # 리뷰 페이지
        ttk.Label(main_frame, text="리뷰 페이지:").grid(row=3, column=0, sticky=tk.W, pady=5)
        ttk.Radiobutton(main_frame, text="5", variable=self.page_var, value="5").grid(row=3, column=1, pady=5)
        ttk.Radiobutton(main_frame, text="15", variable=self.page_var, value="15").grid(row=3, column=2, pady=5)
        ttk.Radiobutton(main_frame, text="50", variable=self.page_var, value="50").grid(row=3, column=3, pady=5)
        ttk.Radiobutton(main_frame, text="max", variable=self.page_var, value="max").grid(row=3, column=4, pady=5)
        
        # 직접 입력
        ttk.Label(main_frame, text="직접 입력:").grid(row=4, column=0, sticky=tk.W, pady=5)
        self.custom_page_entry = ttk.Entry(main_frame, textvariable=self.custom_page_var, width=10)
        self.custom_page_entry.grid(row=4, column=1, sticky=tk.W, pady=5)
        self.custom_page_entry.bind('<KeyRelease>', self.on_custom_page_change)
        
        # 저장 경로
        ttk.Label(main_frame, text="저장 경로:").grid(row=5, column=0, sticky=tk.W, pady=5)
        self.path_input = ttk.Entry(main_frame, width=40)
        self.path_input.grid(row=5, column=1, columnspan=2, sticky=(tk.W, tk.E), pady=5)
        ttk.Button(main_frame, text="선택", command=self.select_path).grid(row=5, column=3, pady=5)
        
        # 폴더 자동 열기 옵션
        ttk.Checkbutton(main_frame, text="크롤링 후 폴더 열기", variable=self.open_folder_var).grid(row=6, column=0, columnspan=4, sticky=tk.W, pady=5)
        
        # 실행 버튼
        self.run_button = ttk.Button(main_frame, text="실행", command=self.run_crawler)
        self.run_button.grid(row=7, column=0, columnspan=4, pady=10)
        
        # 로그 출력
        ttk.Label(main_frame, text="로그:").grid(row=8, column=0, sticky=tk.W, pady=5)
        self.log_text = scrolledtext.ScrolledText(main_frame, height=15, width=60)
        self.log_text.grid(row=9, column=0, columnspan=4, sticky=(tk.W, tk.E, tk.N, tk.S), pady=5)
        
        # 그리드 가중치 설정
        main_frame.columnconfigure(1, weight=1)
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)

    def on_custom_page_change(self, event):
        if self.custom_page_var.get():
            self.page_var.set("custom")

    def select_path(self):
        folder_path = filedialog.askdirectory(title="저장 경로 선택")
        if folder_path:
            self.path_input.delete(0, tk.END)
            self.path_input.insert(0, folder_path)

    def update_log(self, message):
        self.log_text.insert(tk.END, message + "\n")
        self.log_text.see(tk.END)
        self.root.update_idletasks()

    def run_crawler(self):
        # 입력값 가져오기
        input_url = self.url_input.get()
        if not input_url:
            messagebox.showerror("오류", "상품 링크를 입력해주세요.")
            return
            
        sort_option = self.sort_var.get()
        
        if self.page_var.get() == "custom":
            input_pages = self.custom_page_var.get()
            if not input_pages:
                messagebox.showerror("오류", "페이지 수를 입력해주세요.")
                return
        else:
            input_pages = self.page_var.get()

        photo_folder_path = self.path_input.get()
        if not photo_folder_path:
            messagebox.showerror("오류", "저장 경로를 선택해주세요.")
            return

        # 크롤링 스레드 시작
        self.crawler_thread = CrawlerThread(input_url, input_pages, photo_folder_path, sort_option, callback=self.update_log)
        self.crawler_thread.daemon = True
        self.crawler_thread.start()

        # UI 비활성화
        self.run_button.config(state='disabled')
        self.update_log("크롤링을 시작합니다...")
        
        # 완료 체크를 위한 타이머
        self.check_thread_completion()

    def check_thread_completion(self):
        if hasattr(self, 'crawler_thread') and self.crawler_thread.is_alive():
            self.root.after(1000, self.check_thread_completion)
        else:
            self.run_button.config(state='normal')
            self.update_log("크롤링이 완료되었습니다.")
            
            # 폴더 자동 열기
            if self.open_folder_var.get():
                folder_path = self.path_input.get()
                if os.path.exists(folder_path):
                    if sys.platform == 'win32':
                        os.startfile(folder_path)
                    elif sys.platform == 'darwin':  # macOS
                        subprocess.Popen(['open', folder_path])
                    else:  # linux
                        subprocess.Popen(['xdg-open', folder_path])

if __name__ == '__main__':
    root = tk.Tk()
    app = ReviewCrawlerGUI(root)
    root.mainloop()